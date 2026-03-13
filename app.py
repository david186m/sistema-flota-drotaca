import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, timedelta
import plotly.express as px
import os
import io
import openpyxl
from fpdf import FPDF
import tempfile
import time 
import streamlit.components.v1 as components 

# --- IMPORTACIÓN DEL NUEVO MÓDULO MODULARIZADO ---
from compras import renderizar_modulo_compras

# --- 1. CONFIGURACIÓN DE PÁGINA (Debe ser la primera línea) ---
st.set_page_config(page_title="Control de Flota Drotaca", page_icon="🚛", layout="wide")

# --- 2. CONFIGURACIÓN DE LA MEMORIA (SESSION STATE) ---
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario_actual" not in st.session_state:
    st.session_state.usuario_actual = ""
if "nombre_real" not in st.session_state:
    st.session_state.nombre_real = ""

# =========================================================================================
# --- VARIABLE GLOBAL DE ESTILOS HTML PARA TODAS LAS TABLAS ---
# =========================================================================================
estilos_html_genericos = [
    dict(selector="table", props=[("width", "100%"), ("border-collapse", "collapse"), ("font-family", "sans-serif"), ("border", "1px solid black"), ("background-color", "white")]),
    dict(selector="thead th", props=[("background-color", "#1A3B5C"), ("color", "white"), ("font-weight", "bold"), ("text-align", "center"), ("padding", "12px"), ("border", "1px solid black"), ("font-size", "14px")]),
    dict(selector="tbody td", props=[("border", "1px solid black")])
]

# --- 3. FUNCIONES GENERALES ---
def obtener_hora_venezuela():
    return datetime.utcnow() - timedelta(hours=4)

def limpiar_numero_logistica(valor):
    if valor is None or valor == "" or str(valor).lower() == "none":
        return 0.0
    s = str(valor).upper().replace('KMS', '').strip()
    if '.' in s and ',' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif '.' in s:
        if len(s.split('.')[-1]) == 3:
            s = s.replace('.', '')
    elif ',' in s:
        if len(s.split(',')[-1]) == 3:
            s = s.replace(',', '')
        else:
            s = s.replace(',', '.')
    try:
        return float(s)
    except:
        return 0.0

# DICCIONARIO PARA MESES EN ESPAÑOL
MESES_ESPANOL = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

# --- 3.1 LIMPIEZA EXTREMA PARA PDF (EVITAR SÍMBOLOS "?" POR EMOJIS) ---
def limpiar_texto_pdf(texto):
    if pd.isna(texto): return ""
    texto = str(texto)
    reemplazos = {
        "Ñ":"N", "ñ":"n", "á":"a", "é":"e", "í":"i", "ó":"o", "ú":"u",
        "Á":"A", "É":"E", "Í":"I", "Ó":"O", "Ú":"U",
        '\u2013':'-', '\u2014':'-', '\u2018':"'", '\u2019':"'", '\u201c':'"', '\u201d':'"', '\u2026':'...'
    }
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
        
    return texto.encode('latin-1', 'ignore').decode('latin-1').strip()

# --- MOTORES GENERADORES DE PDF ---
def crear_pdf_operativo(nombre, rol, df_datos, dias_tot, dias_trab, dias_inac, mes_filtro, extra_filtros=""):
    class PDF(FPDF):
        def header(self):
            if os.path.exists("encabezado.png"):
                self.image("encabezado.png", x=10, y=8, w=277)
                self.set_y(46)
            else:
                self.set_y(15)
                
            self.set_font('Arial', 'B', 14)
            self.set_fill_color(26, 59, 92)
            self.set_text_color(255, 255, 255)
            self.cell(0, 10, f' DROTACA - REPORTE DE ROTACION: {rol.upper()}', 0, 1, 'C', 1)
            self.ln(3)

    pdf = PDF('L', 'mm', 'A4') 
    pdf.add_page()
    
    pdf.set_font('Arial', 'B', 12)
    pdf.set_text_color(0, 0, 0)
    nombre_safe = limpiar_texto_pdf(nombre)
    pdf.cell(0, 6, f"Perfil Operativo: {nombre_safe}", 0, 1)
    
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Periodo: {limpiar_texto_pdf(mes_filtro)} {limpiar_texto_pdf(extra_filtros)}", 0, 1)

    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 6, f"Dias Registrados: {dias_tot}  |  Trabajados: {dias_trab}  |  Inactivos: {dias_inac}", 0, 1)
    pdf.ln(4)

    columnas = df_datos.columns.tolist()
    
    if rol == 'Chofer':
        anchos = [18, 18, 33, 119, 22, 65] 
    else:
        anchos = [18, 18, 38, 30, 90, 20, 61] 

    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(230, 230, 230)
    for i, col in enumerate(columnas):
        pdf.cell(anchos[i], 8, limpiar_texto_pdf(col), 1, 0, 'C', 1)
    pdf.ln()

    pdf.set_font('Arial', '', 8)
    for index, row in df_datos.iterrows():
        for i, col in enumerate(columnas):
            texto = limpiar_texto_pdf(row[col])
            max_chars = int(anchos[i] * 0.70) 
            texto_limpio = texto[:max_chars]
            pdf.cell(anchos[i], 7, texto_limpio, 1, 0, 'L')
        pdf.ln()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    os.remove(tmp.name)
    return data

def crear_pdf_novedades(df_datos, texto_filtros):
    class PDFNovedades(FPDF):
        def header(self):
            if os.path.exists("encabezado.png"):
                self.image("encabezado.png", x=10, y=8, w=277)
                self.set_y(46)
            else:
                self.set_y(15)
                
            self.set_font('Arial', 'B', 15)
            self.set_fill_color(26, 59, 92)
            self.set_text_color(255, 255, 255)
            self.cell(0, 12, ' DROTACA - BITACORA DE NOVEDADES EN RUTA', 0, 1, 'C', 1)
            self.ln(4)

    pdf = PDFNovedades('L', 'mm', 'A4') 
    pdf.add_page()
    
    pdf.set_font('Arial', 'B', 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, f"Filtros Aplicados: {limpiar_texto_pdf(texto_filtros)}", 0, 1)
    pdf.cell(0, 6, f"Total Registros: {len(df_datos)}", 0, 1)
    pdf.ln(4)

    columnas_principales = ['FECHA', 'HORA', 'RUTA', 'ZONA', 'PLACA', 'UNIDAD', 'CHOFER', 'AYUDANTE', 'TIPO DE NOVEDAD']
    anchos = [18, 12, 45, 18, 18, 20, 35, 35, 76] 
    
    pdf.set_font('Arial', 'B', 8)
    pdf.set_fill_color(230, 230, 230)
    pdf.set_text_color(0, 0, 0)
    for i, col in enumerate(columnas_principales):
        pdf.cell(anchos[i], 8, limpiar_texto_pdf(col), 1, 0, 'C', 1)
    pdf.ln()

    for index, row in df_datos.iterrows():
        pdf.set_font('Arial', 'B', 7)
        pdf.set_text_color(0, 0, 0)
        for i, col in enumerate(columnas_principales):
            texto = limpiar_texto_pdf(row[col])
            max_chars = int(anchos[i] * 0.60) 
            texto_limpio = texto[:max_chars]
            alineacion = 'C' if col in ['FECHA', 'HORA', 'PLACA', 'ZONA'] else 'L'
            pdf.cell(anchos[i], 6, texto_limpio, 1, 0, alineacion)
        pdf.ln()
        
        desc_texto = limpiar_texto_pdf(row.get('DESCRIPCIÓN', 'Sin descripción detallada.'))
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(50, 50, 50)
        pdf.multi_cell(277, 5, f"Observacion Detallada: {desc_texto}", border=1, align='L', fill=False)
        pdf.ln(2) 

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    os.remove(tmp.name)
    return data

def crear_pdf_taller(df_datos, texto_filtros):
    class PDFTaller(FPDF):
        def header(self):
            if os.path.exists("encabezado.png"):
                self.image("encabezado.png", x=10, y=8, w=277)
                self.set_y(46)
            else:
                self.set_y(15)
                
            self.set_font('Arial', 'B', 15)
            self.set_fill_color(26, 59, 92)
            self.set_text_color(255, 255, 255)
            self.cell(0, 12, ' DROTACA - HISTORIAL DE MANTENIMIENTO TALLER', 0, 1, 'C', 1)
            self.ln(4)

    pdf = PDFTaller('L', 'mm', 'A4') 
    pdf.add_page()
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Filtros: {limpiar_texto_pdf(texto_filtros)}", 0, 1)
    pdf.cell(0, 5, f"Total Registros: {len(df_datos)}", 0, 1)
    pdf.ln(3)

    titulos_pdf = ['PLACA', 'RUTA', 'ZONA', 'ENTRADA', 'SALIDA', 'FALLA', 'MECANICO', 'DURACION', 'ESTATUS']
    anchos      = [  15,      45,     15,       18,        18,      83,         35,         18,        30   ] 
    
    pdf.set_font('Arial', 'B', 8)
    pdf.set_fill_color(230, 230, 230)
    pdf.set_text_color(0, 0, 0)
    for i, col in enumerate(titulos_pdf):
        pdf.cell(anchos[i], 8, limpiar_texto_pdf(col), 1, 0, 'C', 1)
    pdf.ln()

    columnas_df = ['Placa', 'Ruta', 'Zona', 'Fecha_Entrada', 'Fecha_Salida', 'Motivo_Falla', 'Taller / Mecánico', 'Duración', 'Estatus_Reparacion']
    
    for index, row in df_datos.iterrows():
        pdf.set_font('Arial', 'B', 7)
        pdf.set_text_color(0, 0, 0)
        
        estado_str = str(row.get('Estatus_Reparacion', '')).strip().upper()
        if 'TERMINADO' in estado_str: pdf.set_text_color(25, 135, 84)
        elif 'EN PROCESO' in estado_str: pdf.set_text_color(220, 53, 69)
        elif '⚠️' in estado_str: pdf.set_text_color(220, 53, 69)
        else: pdf.set_text_color(0, 0, 0)
        
        if '⚠️' in str(row.get('Duración', '')): pdf.set_text_color(220, 53, 69)

        for i, col in enumerate(columnas_df):
            val = str(row.get(col, ''))
            texto = limpiar_texto_pdf(val)
            alineacion = 'C' if col in ['Placa', 'Zona', 'Fecha_Entrada', 'Fecha_Salida', 'Duración', 'Estatus_Reparacion'] else 'L'
            
            texto_cortado = texto[:int(anchos[i] * 0.62)]
            pdf.cell(anchos[i], 6, texto_cortado, 1, 0, alineacion)
        pdf.ln()
        
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    os.remove(tmp.name)
    return data

# --- 4. PROCESAMIENTO DE DATOS (MÓDULO DE FLOTA) ---
@st.cache_data(ttl=60)
def cargar_y_procesar_datos():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    cliente = gspread.authorize(creds)
    libro_flota = cliente.open("Sistema_Flota_2026")
    ws_km = libro_flota.worksheet("Kilometraje")
    ws_control = libro_flota.worksheet("Control_Diario")
    ws_maestro = libro_flota.worksheet("Maestro_Flota")
    ws_taller = libro_flota.worksheet("Historial_Taller")
    
    try:
        hora_sincronizacion = ws_control.acell('Z2').value
        if not hora_sincronizacion:
            hora_sincronizacion = "Esperando primera edición..."
    except:
        hora_sincronizacion = "Sin registro en celda Z2"

    df_km = pd.DataFrame(ws_km.get_all_values()[1:], columns=ws_km.get_all_values()[0])
    df_control = pd.DataFrame(ws_control.get_all_values()[1:], columns=ws_control.get_all_values()[0])
    df_maestro = pd.DataFrame(ws_maestro.get_all_values()[1:], columns=ws_maestro.get_all_values()[0])
    
    datos_taller = ws_taller.get_all_values()
    columnas_taller = ['Placa', 'Ruta', 'Zona', 'Fecha_Entrada', 'Motivo_Falla', 'Estatus_Reparacion', 'Fecha_Salida', 'Taller / Mecánico']
    df_taller = pd.DataFrame(datos_taller[1:], columns=datos_taller[0]) if len(datos_taller) > 1 else pd.DataFrame(columns=columnas_taller)
    
    if not df_taller.empty and 'Fecha_Entrada' in df_taller.columns:
        df_taller['Fecha_Entrada_DT'] = pd.to_datetime(df_taller['Fecha_Entrada'], dayfirst=True, errors='coerce')
        df_taller['Fecha_Salida_DT'] = pd.to_datetime(df_taller['Fecha_Salida'], dayfirst=True, errors='coerce')
        hoy = pd.Timestamp.today().normalize()
        def calcular_duracion(row):
            if pd.isna(row['Fecha_Entrada_DT']): return ""
            fecha_fin = row['Fecha_Salida_DT'] if pd.notna(row['Fecha_Salida_DT']) else hoy
            dias = max(0, (fecha_fin - row['Fecha_Entrada_DT']).days)
            return f"⚠️ {dias} días" if dias > 10 else f"{dias} días"
        df_taller['Duración'] = df_taller.apply(calcular_duracion, axis=1)
        df_taller = df_taller.drop(columns=['Fecha_Entrada_DT', 'Fecha_Salida_DT'])
        if 'Duración' in df_taller.columns and 'Fecha_Salida' in df_taller.columns:
            cols = df_taller.columns.tolist()
            cols.insert(cols.index('Fecha_Salida') + 1, cols.pop(cols.index('Duración')))
            df_taller = df_taller[cols]

    df_km['FECHA_DT'] = pd.to_datetime(df_km['FECHA'], format='%d/%m/%Y', errors='coerce')
    df_km['KM_LIMPIO'] = df_km['KILOMETRAJE'].apply(limpiar_numero_logistica)
    df_validos = df_km[df_km['KM_LIMPIO'] > 0].copy()
    df_ultimo = df_validos.sort_values('FECHA_DT', ascending=False).drop_duplicates('UNIDAD')
    inicio_mes = pd.Timestamp(2026, 3, 1)
    
    km_historico = df_validos.groupby(['UNIDAD', 'FECHA_DT'])['KM_LIMPIO'].max().reset_index()
    km_historico = km_historico.sort_values(['UNIDAD', 'FECHA_DT'])
    km_historico['Recorrido_Dia'] = km_historico.groupby('UNIDAD')['KM_LIMPIO'].diff().fillna(0)
    km_mes = km_historico[km_historico['FECHA_DT'] >= inicio_mes]
    dias_activos_df = km_mes[km_mes['Recorrido_Dia'] >= 70].groupby('UNIDAD').size().reset_index(name='dias_activos')
    
    df_mes_actual = df_validos[df_validos['FECHA_DT'] >= inicio_mes]
    df_mes_anterior = df_validos[df_validos['FECHA_DT'] < inicio_mes]
    recorrido_mes = df_mes_actual.groupby('UNIDAD').agg(km_max_mes=('KM_LIMPIO', 'max'), km_min_mes_actual=('KM_LIMPIO', 'min')).reset_index()
    cierre_mes_anterior = df_mes_anterior.groupby('UNIDAD').agg(km_cierre_anterior=('KM_LIMPIO', 'max')).reset_index()
    recorrido_mes = pd.merge(recorrido_mes, cierre_mes_anterior, on='UNIDAD', how='left')
    recorrido_mes['km_arranque'] = recorrido_mes['km_cierre_anterior'].fillna(recorrido_mes['km_min_mes_actual'])
    recorrido_mes['Km Mensual Actual'] = recorrido_mes['km_max_mes'] - recorrido_mes['km_arranque']
    recorrido_mes = pd.merge(recorrido_mes, dias_activos_df, on='UNIDAD', how='left')
    recorrido_mes['dias_activos'] = recorrido_mes['dias_activos'].fillna(0).apply(lambda x: x if x > 0 else 1)

    df_merge = pd.merge(df_control[['Placa', 'Grupo', 'RUTA']], df_ultimo[['UNIDAD', 'KM_LIMPIO', 'FECHA']], left_on="Placa", right_on="UNIDAD", how="left")
    df_merge2 = pd.merge(df_merge, recorrido_mes[['UNIDAD', 'Km Mensual Actual', 'dias_activos']], on="UNIDAD", how="left")
    
    if 'Placa' in df_maestro.columns:
        cols_maestro = ['Placa']
        if 'Fecha_GPS' in df_maestro.columns: cols_maestro.append('Fecha_GPS')
        if 'Modelo' in df_maestro.columns: cols_maestro.append('Modelo')
        df_final = pd.merge(df_merge2, df_maestro[cols_maestro], on="Placa", how="left")
        if 'Fecha_GPS' in df_final.columns:
            df_final['Estatus GPS'] = df_final['Fecha_GPS'].apply(lambda x: 'Sin GPS' if pd.isna(x) or str(x).strip().upper() == 'SIN GPS' or str(x).strip() == '' else 'GPS Operativo')
            df_final = df_final.drop(columns=['Fecha_GPS'])
        if 'Modelo' in df_final.columns:
            df_final['Modelo'] = df_final['Modelo'].replace('', 'N/A').fillna('N/A')
    else:
        df_final = df_merge2
        df_final['Estatus GPS'] = "Error"
        df_final['Modelo'] = "Error"
    
    if not df_taller.empty and 'Placa' in df_taller.columns and 'Estatus_Reparacion' in df_taller.columns:
        df_taller_ultimo = df_taller.drop_duplicates(subset=['Placa'], keep='last').copy()
        en_taller = df_taller_ultimo['Estatus_Reparacion'].str.strip().str.upper() != 'TERMINADO'
        df_activos = df_taller_ultimo[en_taller][['Placa', 'Fecha_Entrada', 'Motivo_Falla', 'Taller / Mecánico']]
        df_final = pd.merge(df_final, df_activos, on="Placa", how="left")
        df_final['Estatus_Unidad'] = df_final['Fecha_Entrada'].apply(lambda x: 'No Operativo' if pd.notna(x) else 'Operativo')
        def armar_observacion(row):
            if pd.notna(row['Motivo_Falla']):
                mecanico = row['Taller / Mecánico'] if pd.notna(row['Taller / Mecánico']) and str(row['Taller / Mecánico']).strip() != '' else 'Sin asignar'
                return f"{row['Motivo_Falla']} ({mecanico})"
            return ""
        df_final['Observacion'] = df_final.apply(armar_observacion, axis=1)
        df_final['Fecha_Inoperativo'] = df_final['Fecha_Entrada'].fillna('')
        df_final = df_final.drop(columns=['Fecha_Entrada', 'Motivo_Falla', 'Taller / Mecánico'])
    else:
        df_final['Estatus_Unidad'] = 'Operativo'
        df_final['Observacion'] = ''
        df_final['Fecha_Inoperativo'] = ''

    df_final = df_final.drop(columns=['UNIDAD'])
    df_final = df_final.rename(columns={'KM_LIMPIO': 'Km Actual', 'FECHA': 'Última Actualización'})
    df_final['Km Actual'] = df_final['Km Actual'].fillna(0)
    df_final['Km Mensual Actual'] = df_final['Km Mensual Actual'].fillna(0)
    df_final['dias_activos'] = df_final['dias_activos'].fillna(1)
    df_final['Última Actualización'] = df_final['Última Actualización'].fillna("Sin Registro")
    
    columnas_orden = ['Placa', 'Modelo', 'Grupo', 'RUTA', 'Estatus_Unidad', 'Estatus GPS', 'Km Actual', 'Km Mensual Actual', 'dias_activos', 'Última Actualización', 'Observacion', 'Fecha_Inoperativo']
    return df_final[columnas_orden], df_taller, hora_sincronizacion

# --- 4.1 PROCESAMIENTO DE DATOS (MÓDULO DE PERSONAL) ---
@st.cache_data(ttl=120)
def cargar_datos_personal():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    cliente = gspread.authorize(creds)
    libro = cliente.open("Sistema_Flota_2026")
    
    try:
        ws_choferes = libro.worksheet("Rotacion_Choferes")
        df_choferes = pd.DataFrame(ws_choferes.get_all_values()[1:], columns=ws_choferes.get_all_values()[0])
        df_choferes['FECHA_DT'] = pd.to_datetime(df_choferes['FECHA'], format='%d/%m/%Y', errors='coerce')
    except:
        df_choferes = pd.DataFrame()

    try:
        ws_ayudantes = libro.worksheet("Rotacion_Ayudantes")
        df_ayudantes = pd.DataFrame(ws_ayudantes.get_all_values()[1:], columns=ws_ayudantes.get_all_values()[0])
        df_ayudantes['FECHA_DT'] = pd.to_datetime(df_ayudantes['FECHA'], format='%d/%m/%Y', errors='coerce')
    except:
        df_ayudantes = pd.DataFrame()
        
    return df_choferes, df_ayudantes

# --- 5. PANTALLA DE LOGIN ---
def pantalla_login():
    st.markdown("""
    <style>
    [data-testid="stApp"] { background: radial-gradient(circle at center, #151b26 0%, #080a0e 100%) !important; color: #ffffff; }
    [data-testid="stHeader"] { background-color: transparent !important; }
    [data-testid="stForm"] { background-color: rgba(28, 34, 45, 0.8) !important; border-radius: 20px !important; border: 1px solid rgba(0, 212, 255, 0.2) !important; box-shadow: 0 15px 35px rgba(0, 0, 0, 0.5) !important; backdrop-filter: blur(10px) !important; padding: 40px 30px !important; }
    
    /* ========================================================================= */
    /* BLINDAJE DE COLORES PARA INPUTS (Soluciona el error en celulares/Dark Mode) */
    /* ========================================================================= */
    [data-testid="stForm"] div[data-baseweb="input"] > div { 
        background-color: #ffffff !important; 
        border: 2px solid #a0a0a0 !important; 
        border-radius: 8px !important; 
        transition: all 0.3s ease !important; 
    }
    [data-testid="stForm"] div[data-baseweb="input"] > div:focus-within { 
        border-color: #00d4ff !important; 
        box-shadow: 0 0 8px rgba(0, 212, 255, 0.5) !important; 
    }
    [data-testid="stForm"] input { 
        color: #000000 !important; 
        -webkit-text-fill-color: #000000 !important; 
        font-weight: bold !important; 
        background-color: transparent !important; 
    }
    /* ========================================================================= */

    [data-testid="stForm"] label p { color: #a0a0a0 !important; font-size: 0.9rem !important; }
    [data-testid="stFormSubmitButton"] button { background: linear-gradient(45deg, #0056b3, #00d4ff) !important; border: none !important; border-radius: 8px !important; color: white !important; font-weight: bold !important; letter-spacing: 1px !important; transition: transform 0.2s, box-shadow 0.2s !important; width: 100% !important; margin-top: 10px !important; }
    [data-testid="stFormSubmitButton"] button:hover { transform: translateY(-2px) !important; box-shadow: 0 5px 15px rgba(0, 212, 255, 0.4) !important; }
    .login-title { color: #ffffff; font-size: 2.2rem; letter-spacing: 2px; text-align: center; margin-bottom: 5px; font-weight: 600; }
    .login-title span { color: #00d4ff; font-weight: 300; }
    .login-subtitle { color: #a0a0a0; font-size: 0.95rem; text-align: center; margin-bottom: 30px; }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<br><br>", unsafe_allow_html=True) 
    
    usuarios_permitidos = {
        "David_Admin": "Drotaca2026",
        "Supervisor_Oriente": "Oriente27",
        "Supervisor_Centro": "25centro",
        "Supervisor_Occidente": "Occidente26",
        "Jsuarez": "295377886",
        "Franluis_pulve": "456789"
    }

    nombres_reales = {
        "David_Admin": "David Mujica",
        "Supervisor_Oriente": "Javier Hidalgo",
        "Supervisor_Centro": "Faisal Yordi",
        "Supervisor_Occidente": "Wiliams Castillo",
        "Jsuarez": "Jose Suarez",
        "Franluis_pulve": "Franluis Pulve"
    }

    col1, col2, col3 = st.columns([1, 1.5, 1])
    with col2:
        nombre_imagen = "logo.png" 
        if os.path.exists(nombre_imagen):
            st.image(nombre_imagen, width=300)
            st.markdown("<br>", unsafe_allow_html=True)

        st.markdown("<div class='login-title'>🚛 Monitoreo De Flota <span>2026</span></div>", unsafe_allow_html=True)
        st.markdown("<div class='login-subtitle'>Acceso Restringido al Sistema</div><br>", unsafe_allow_html=True)

        with st.form("formulario_login"):
            usuario_input = st.text_input("👤 Usuario:")
            password_input = st.text_input("🔑 Contraseña:", type="password")
            boton_ingresar = st.form_submit_button("Ingresar al Sistema", use_container_width=True)
            
            if boton_ingresar:
                if usuario_input in usuarios_permitidos and usuarios_permitidos[usuario_input] == password_input:
                    with st.spinner("⏳ Verificando credenciales y registrando acceso..."):
                        st.session_state.autenticado = True
                        st.session_state.usuario_actual = usuario_input
                        st.session_state.nombre_real = nombres_reales.get(usuario_input, usuario_input)
                        
                        try:
                            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
                            try:
                                credenciales_dict = dict(st.secrets["gcp_service_account"])
                                creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
                            except:
                                creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
                            cliente = gspread.authorize(creds)
                            libro = cliente.open("Sistema_Flota_2026")
                            
                            try:
                                ws_log = libro.worksheet("Registro_Accesos")
                                ahora = obtener_hora_venezuela()
                                ws_log.append_row([ahora.strftime("%d/%m/%Y"), ahora.strftime("%I:%M %p"), st.session_state.nombre_real])
                            except Exception:
                                pass 
                        except Exception:
                            pass 

                        st.rerun() 
                else:
                    st.error("❌ Usuario o contraseña incorrectos.")

# --- 6. MÓDULO 1: CONTROL DE FLOTA ---
def modulo_flota():
    col_titulo, col_boton = st.columns([0.8, 0.2])
    with col_titulo:
        st.title("🚛 Panel de Control de Flota - Drotaca")
    with col_boton:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Datos", key="btn_flota", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    st.divider()

    try:
        df_resultados, df_historial_taller, ultima_sync = cargar_y_procesar_datos()
        ahora = obtener_hora_venezuela()
        dias_semana = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        meses_año = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        fecha_reloj = f"{dias_semana[ahora.weekday()]}, {ahora.strftime('%d-%m-%Y - %I:%M %p')}"
        mes_actual_texto = f"{meses_año[ahora.month - 1]} {ahora.year}"
        
        st.markdown(f"""
        <div style="display: flex; justify-content: space-between; flex-wrap: wrap; background-color: #f8f9fa; padding: 15px; border-radius: 8px; border-left: 5px solid #003366; margin-bottom: 20px;">
            <div style="font-size: 15px; margin-bottom: 5px;">
                🕒 <b>Hora del Sistema:</b> <span style="color: #003366;">{fecha_reloj}</span>
            </div>
            <div style="font-size: 15px;">
                📡 <b>Última edición real del documento:</b> <span style="color: #198754; font-weight: bold;">{ultima_sync}</span> <span style="font-size: 13px; color: gray;">(Por el equipo)</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        tab1, tab2 = st.tabs(["📊 Monitoreo en Tiempo Real", "🛠️ Gestión de Taller"])
        
        with tab1:
            opciones = ["Todos los vehículos"] + sorted(df_resultados["Grupo"].unique().tolist())
            seleccion = st.selectbox("📌 Filtrar por Región/Grupo:", opciones)
            
            df_mostrar = df_resultados if seleccion == "Todos los vehículos" else df_resultados[df_resultados["Grupo"] == seleccion]
            
            total_unidades = len(df_mostrar)
            total_km_mensual = df_mostrar['Km Mensual Actual'].sum()
            sin_gps_count = len(df_mostrar[df_mostrar['Estatus GPS'] == 'Sin GPS'])
            no_operativas_count = len(df_mostrar[df_mostrar['Estatus_Unidad'] == 'No Operativo'])
            operativas_count = total_unidades - no_operativas_count

            st.markdown("<br>", unsafe_allow_html=True)
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1: st.metric(label="Total Flota", value=f"{total_unidades} 🚚")
            with col2: st.metric(label="✅ Operativas", value=f"{operativas_count}")
            with col3: st.metric(label="⚠️ En Taller", value=f"{no_operativas_count}")
            with col4: st.metric(label="❌ Sin GPS", value=f"{sin_gps_count}")
            with col5: st.metric(label="Recorrido Mensual", value=f"{total_km_mensual:,.2f} km")
            st.divider()

            st.subheader(f"📈 Análisis de Recorrido: {mes_actual_texto}")
            if seleccion == "Todos los vehículos":
                km_data = df_mostrar.groupby("Grupo")["Km Mensual Actual"].sum().reset_index()
                eje_x = "Grupo"
            else:
                km_data = df_mostrar[["Placa", "Km Mensual Actual"]].copy()
                eje_x = "Placa"
                
            km_data = km_data.sort_values(by="Km Mensual Actual", ascending=False)
            km_data['Etiqueta'] = km_data['Km Mensual Actual'].apply(lambda x: f"<b>{x:,.0f} Kms</b>".replace(",", "X").replace(".", ",").replace("X", "."))
            fig = px.bar(km_data, x=eje_x, y="Km Mensual Actual", text="Etiqueta")
            fig.update_traces(
                textposition='outside', 
                marker_color='#1A3B5C', 
                cliponaxis=False,
                textfont=dict(size=16, color='black')
            )
            fig.update_layout(
                xaxis_title="", 
                yaxis_title="Kilómetros Recorridos", 
                dragmode=False, 
                margin=dict(t=40, b=0, l=0, r=0),
                height=400
            )
            fig.update_xaxes(tickfont=dict(size=15, color='black', family='Arial Black')) 
            st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False, 'scrollZoom': False})
            st.divider()
            
            df_promedios = df_mostrar.copy()
            df_promedios['Km_Num'] = pd.to_numeric(df_promedios['Km Mensual Actual'], errors='coerce').fillna(0)
            df_promedios['Odometer_Num'] = pd.to_numeric(df_promedios['Km Actual'], errors='coerce').fillna(0)
            df_promedios['Dias_Num'] = pd.to_numeric(df_promedios['dias_activos'], errors='coerce').fillna(1)
            dias_calendario = ahora.day if ahora.day > 0 else 1

            df_promedios['Promedio_Diario_Num'] = df_promedios['Km_Num'] / df_promedios['Dias_Num']
            ritmo_diario_real = df_promedios['Km_Num'] / dias_calendario
            df_promedios['Promedio_Semanal_Num'] = ritmo_diario_real * 7
            df_promedios['Promedio_Mensual_Num'] = df_promedios['Km_Num'] 

            def format_kms(val):
                try:
                    return f"{int(val):,} Kms".replace(',', '.')
                except:
                    return "0 Kms"

            df_promedios['RECORRIDO PROMEDIO DIARIO'] = df_promedios['Promedio_Diario_Num'].apply(format_kms)
            df_promedios['RECORRIDO PROMEDIO SEMANAL'] = df_promedios['Promedio_Semanal_Num'].apply(format_kms)
            df_promedios['RECORRIDO PROMEDIO MENSUAL'] = df_promedios['Promedio_Mensual_Num'].apply(format_kms)
            df_promedios['Km Actual Formato'] = df_promedios['Odometer_Num'].apply(format_kms)

            columnas_finales = [
                'Placa', 'Modelo', 'Grupo', 'RUTA', 'Estatus_Unidad', 'Estatus GPS', 
                'Km Actual Formato', 'RECORRIDO PROMEDIO DIARIO', 'RECORRIDO PROMEDIO SEMANAL', 
                'RECORRIDO PROMEDIO MENSUAL', 'Última Actualización', 'Observacion', 'Fecha_Inoperativo'
            ]
            
            columnas_existentes = [col for col in columnas_finales if col in df_promedios.columns]
            df_display = df_promedios[columnas_existentes].rename(columns={'Km Actual Formato': 'Km Actual'}).copy()

            col_tabla1, col_tabla2 = st.columns([0.7, 0.3])
            with col_tabla1:
                st.subheader(f"📑 Reporte Detallado: {seleccion}")
            with col_tabla2:
                try:
                    df_export = df_promedios.copy()
                    df_export['RUTA_STR'] = df_export['RUTA'].astype(str).fillna('')
                    df_export = df_export.sort_values(by='RUTA_STR', ascending=True).drop(columns=['RUTA_STR'])
                    
                    wb = openpyxl.load_workbook("INFORME GERENCIAL.xlsx")
                    ws = wb.active 
                    titulo_zona = f"INFORME MENSUAL - RUTA {seleccion.upper()} 2026" if seleccion != "Todos los vehículos" else "INFORME MENSUAL - TODA LA FLOTA 2026"
                    ws['E1'] = titulo_zona
                    ws['J1'] = meses_año[ahora.month - 1].upper()

                    fila_inicio = 3
                    for index, row in enumerate(df_export.to_dict('records')):
                        fila_actual = fila_inicio + index
                        ws.cell(row=fila_actual, column=1, value=index + 1)
                        ws.cell(row=fila_actual, column=2, value=row.get('Placa', ''))
                        ws.cell(row=fila_actual, column=3, value=row.get('Modelo', ''))
                        ws.cell(row=fila_actual, column=4, value=row.get('Grupo', ''))
                        ws.cell(row=fila_actual, column=5, value=row.get('RUTA', ''))
                        ws.cell(row=fila_actual, column=6, value=row.get('Estatus_Unidad', ''))
                        ws.cell(row=fila_actual, column=7, value=row.get('Estatus GPS', ''))
                        ws.cell(row=fila_actual, column=8, value=row.get('Odometer_Num', 0))
                        ws.cell(row=fila_actual, column=9, value=row.get('Promedio_Diario_Num', 0))
                        ws.cell(row=fila_actual, column=10, value=row.get('Promedio_Semanal_Num', 0))
                        ws.cell(row=fila_actual, column=11, value=row.get('Promedio_Mensual_Num', 0)) 

                    for r in range(fila_inicio + len(df_export), 72):
                        ws.row_dimensions[r].hidden = True

                    output = io.BytesIO()
                    wb.save(output)
                    output.seek(0)

                    st.download_button(
                        label="📥 Descargar Informe Gerencial",
                        data=output,
                        file_name=f"INFORME_GERENCIAL_{seleccion.replace(' ', '_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except FileNotFoundError:
                    st.warning("⚠️ Sube 'INFORME GERENCIAL.xlsx' a tu carpeta.")
            
            total_diario = df_promedios['Promedio_Diario_Num'].sum()
            total_semanal = df_promedios['Promedio_Semanal_Num'].sum()
            total_mensual = df_promedios['Promedio_Mensual_Num'].sum()

            totals_row = {col: "" for col in df_display.columns}
            totals_row['Placa'] = "TOTALES"
            if 'RECORRIDO PROMEDIO DIARIO' in totals_row: totals_row['RECORRIDO PROMEDIO DIARIO'] = format_kms(total_diario)
            if 'RECORRIDO PROMEDIO SEMANAL' in totals_row: totals_row['RECORRIDO PROMEDIO SEMANAL'] = format_kms(total_semanal)
            if 'RECORRIDO PROMEDIO MENSUAL' in totals_row: totals_row['RECORRIDO PROMEDIO MENSUAL'] = format_kms(total_mensual)

            df_display = pd.concat([df_display, pd.DataFrame([totals_row])], ignore_index=True)

            def aplicar_estilos_html_flota(row):
                styles = [''] * len(row)
                if row['Placa'] == 'TOTALES':
                    return ['background-color: #ffff00; color: black; font-weight: bold; text-align: center; border: 1px solid black; padding: 10px; font-size: 13px;'] * len(row)
                
                for i, col in enumerate(row.index):
                    base_style = 'border: 1px solid black; text-align: center; padding: 10px; font-size: 13px; color: black; background-color: white;'
                    
                    if col == 'Estatus GPS':
                        color_t = '#198754' if row[col] == 'GPS Operativo' else '#DC3545'
                        styles[i] = base_style + f' color: {color_t}; font-weight: bold;'
                    elif col == 'Estatus_Unidad':
                        val_str = str(row[col]).strip().upper()
                        if val_str == 'OPERATIVO': styles[i] = base_style + ' color: #198754; font-weight: bold;'
                        elif val_str == 'NO OPERATIVO': styles[i] = base_style + ' color: #DC3545; font-weight: bold;'
                        else: styles[i] = base_style
                    elif col in ['RECORRIDO PROMEDIO DIARIO', 'RECORRIDO PROMEDIO SEMANAL', 'RECORRIDO PROMEDIO MENSUAL']:
                        bg_color = '#1f497d' if 'DIARIO' in col else ('#4f81bd' if 'SEMANAL' in col else '#e46c0a')
                        styles[i] = base_style + f' background-color: {bg_color}; color: white; font-weight: bold;'
                    else: 
                        styles[i] = base_style
                return styles

            tabla_html_flota = df_display.style.apply(aplicar_estilos_html_flota, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
            st.markdown(tabla_html_flota, unsafe_allow_html=True)

        with tab2:
            st.subheader("🛠️ Registro Histórico de Mantenimiento")
            
            if not df_historial_taller.empty:
                df_taller_view = df_historial_taller.copy()
                if 'Placa' in df_taller_view.columns:
                    df_taller_view = df_taller_view[df_taller_view['Placa'].astype(str).str.strip().str.lower() != 'nan']
                    df_taller_view = df_taller_view[df_taller_view['Placa'].astype(str).str.strip() != '']
                
                if df_taller_view.empty:
                    st.info("No hay registros de taller en el sistema.")
                else:
                    df_taller_view['FECHA_DT'] = pd.to_datetime(df_taller_view['Fecha_Entrada'], dayfirst=True, errors='coerce')
                    df_taller_view['MES_NUM'] = df_taller_view['FECHA_DT'].dt.month
                    df_taller_view['MES_NOMBRE'] = df_taller_view['MES_NUM'].map(MESES_ESPANOL).fillna("Desconocido")
                    df_taller_view['SEMANA'] = df_taller_view['FECHA_DT'].dt.isocalendar().week.astype(str).replace('<NA>', 'Desconocida')

                    st.markdown("#### 🔍 Filtros Avanzados de Taller")
                    col_t1, col_t2, col_t3, col_t4 = st.columns(4)
                    
                    fechas_disp = ["Todas"] + sorted([str(x) for x in df_taller_view['Fecha_Entrada'].unique() if str(x).strip() != ""])
                    meses_disp = ["Todos"] + sorted(list(set([m for m in df_taller_view['MES_NOMBRE'] if m != "Desconocido"])), key=lambda m: list(MESES_ESPANOL.values()).index(m) if m in MESES_ESPANOL.values() else 0)
                    semanas_disp = ["Todas"] + sorted([str(x) for x in df_taller_view['SEMANA'].unique() if x != "Desconocida"], reverse=True)
                    estatus_disp = ["Todos", "En Proceso", "Terminado"]
                    
                    with col_t1: f_taller_mes = st.selectbox("📅 Mes de Ingreso:", meses_disp, key="taller_f_mes")
                    with col_t2: f_taller_sem = st.selectbox("📆 Semana:", semanas_disp, key="taller_f_sem")
                    with col_t3: f_taller_fec = st.selectbox("📌 Fecha Exacta:", fechas_disp, key="taller_f_fec")
                    with col_t4: f_taller_est = st.selectbox("⚙️ Estatus:", estatus_disp, key="taller_f_est")
                    
                    busqueda_taller = st.text_input("🔎 Búsqueda libre (Placa, Palabra Clave o Mecánico. Ej: Motor):", key="taller_busq").strip().upper()

                    filtros_taller = []
                    if f_taller_mes and f_taller_mes not in ["Todos", "Todas"]:
                        df_taller_view = df_taller_view[df_taller_view['MES_NOMBRE'] == f_taller_mes]
                        filtros_taller.append(f"Mes: {f_taller_mes}")
                        
                    if f_taller_sem and f_taller_sem not in ["Todos", "Todas"]:
                        df_taller_view = df_taller_view[df_taller_view['SEMANA'] == f_taller_sem]
                        filtros_taller.append(f"Semana: {f_taller_sem}")
                        
                    if f_taller_fec and f_taller_fec not in ["Todos", "Todas"]:
                        df_taller_view = df_taller_view[df_taller_view['Fecha_Entrada'].astype(str).str.strip() == str(f_taller_fec).strip()]
                        filtros_taller.append(f"Fecha: {f_taller_fec}")
                        
                    if f_taller_est and f_taller_est not in ["Todos", "Todas"]:
                        df_taller_view = df_taller_view[df_taller_view['Estatus_Reparacion'].astype(str).str.strip().str.upper() == str(f_taller_est).strip().upper()]
                        filtros_taller.append(f"Estatus: {f_taller_est}")
                        
                    if busqueda_taller and busqueda_taller != "":
                        mask = df_taller_view.astype(str).apply(lambda x: x.str.contains(busqueda_taller, case=False)).any(axis=1)
                        df_taller_view = df_taller_view[mask]
                        filtros_taller.append(f"Búsqueda: '{busqueda_taller}'")
                        
                    texto_filtros_t = " | ".join(filtros_taller) if filtros_taller else "Ninguno (Mostrando todo)"
                    
                    cols_a_ocultar = ['FECHA_DT', 'MES_NUM', 'MES_NOMBRE', 'SEMANA']
                    df_html_taller = df_taller_view.drop(columns=[c for c in cols_a_ocultar if c in df_taller_view.columns])

                    st.markdown("<br>", unsafe_allow_html=True)
                    col_m1, col_m2 = st.columns([0.7, 0.3])
                    with col_m1:
                        st.metric("Total Registros Encontrados", len(df_html_taller))
                    with col_m2:
                        if not df_html_taller.empty:
                            pdf_taller_bytes = crear_pdf_taller(df_html_taller, texto_filtros_t)
                            st.download_button(
                                label="📄 Descargar Historial en PDF",
                                data=pdf_taller_bytes,
                                file_name=f"Historial_Taller_{obtener_hora_venezuela().strftime('%d%m%Y')}.pdf",
                                mime="application/pdf",
                                use_container_width=True,
                                type="primary"
                            )

                    if not df_html_taller.empty:
                        def aplicar_estilos_html_taller(row):
                            styles = [''] * len(row)
                            for i, col in enumerate(row.index):
                                base_style = 'border: 1px solid black; text-align: center; padding: 10px; font-size: 13px; color: black; background-color: white;'
                                if col == 'Estatus_Reparacion':
                                    val_str = str(row[col]).strip().upper()
                                    if 'TERMINADO' in val_str: styles[i] = base_style + ' color: #198754; font-weight: bold; background-color: #E8F5E9;'
                                    elif 'EN PROCESO' in val_str: styles[i] = base_style + ' color: #DC3545; font-weight: bold; background-color: #FFEFEF;'
                                    elif '⚠️' in val_str: styles[i] = base_style + ' color: #DC3545; font-weight: bold; background-color: #FFF3CD;'
                                    else: styles[i] = base_style
                                elif col == 'Duración' and '⚠️' in str(row[col]):
                                    styles[i] = base_style + ' color: #DC3545; font-weight: bold; background-color: #FFF3CD;'
                                elif col in ['Motivo_Falla', 'Taller / Mecánico']:
                                    styles[i] = base_style + ' text-align: left;'
                                else:
                                    styles[i] = base_style
                            return styles

                        tabla_html_taller = df_html_taller.style.apply(aplicar_estilos_html_taller, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
                        st.markdown(tabla_html_taller, unsafe_allow_html=True)
                    else:
                        st.info("No hay registros que coincidan con los filtros de taller seleccionados.")
            else:
                st.info("No hay registros de taller en el sistema.")

    except Exception as e:
        st.error(f"Error cargando los datos de Flota: {e}")

# --- 7. MÓDULO 2: ROTACIÓN DE PERSONAL ---
def modulo_personal():
    col_titulo, col_boton = st.columns([0.8, 0.2])
    with col_titulo:
        st.title("👥 Control y Rotación de Personal")
    with col_boton:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Datos", key="btn_personal", use_container_width=True):
            cargar_datos_personal.clear()
            st.rerun()
    
    try:
        df_choferes, df_ayudantes = cargar_datos_personal()
        
        if df_choferes.empty and df_ayudantes.empty:
            st.warning("⚠️ No se encontraron las hojas 'Rotacion_Choferes' o 'Rotacion_Ayudantes' en Google Sheets.")
            return

        if not df_choferes.empty:
            df_choferes['MES_NUM'] = df_choferes['FECHA_DT'].dt.month
            df_choferes['MES_NOMBRE'] = df_choferes['MES_NUM'].map(MESES_ESPANOL)
        if not df_ayudantes.empty:
            df_ayudantes['MES_NUM'] = df_ayudantes['FECHA_DT'].dt.month
            df_ayudantes['MES_NOMBRE'] = df_ayudantes['MES_NUM'].map(MESES_ESPANOL)

        meses_disponibles = set()
        if not df_choferes.empty: meses_disponibles.update(df_choferes['MES_NOMBRE'].dropna().unique())
        if not df_ayudantes.empty: meses_disponibles.update(df_ayudantes['MES_NOMBRE'].dropna().unique())
        
        meses_ordenados = sorted(list(meses_disponibles), key=lambda m: list(MESES_ESPANOL.values()).index(m))

        st.markdown("""<div style="background-color: #ffffff; padding: 10px 20px; border-radius: 8px; border-left: 5px solid #1A3B5C; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.05);"></div>""", unsafe_allow_html=True)
        col_mes1, col_mes2 = st.columns([0.3, 0.7])
        with col_mes1:
            mes_seleccionado = st.selectbox("📅 Filtrar por Mes:", ["Todo el año"] + meses_ordenados)

        if mes_seleccionado != "Todo el año":
            if not df_choferes.empty: df_choferes = df_choferes[df_choferes['MES_NOMBRE'] == mes_seleccionado]
            if not df_ayudantes.empty: df_ayudantes = df_ayudantes[df_ayudantes['MES_NOMBRE'] == mes_seleccionado]

        st.divider()

        def estilo_personal_html(row):
            styles = [''] * len(row)
            for i, col in enumerate(row.index):
                base_style = 'border: 1px solid black; text-align: center; padding: 10px; font-size: 13px; color: black; background-color: white;'
                if col == 'OBSERVACIÓN':
                    styles[i] = base_style + ' text-align: left; font-style: italic;'
                else:
                    styles[i] = base_style
            return styles

        tab_choferes, tab_ayudantes = st.tabs(["🚛 Gestión de Choferes", "👷 Gestión de Ayudantes"])

        with tab_choferes:
            if not df_choferes.empty:
                lista_choferes = sorted([str(x) for x in df_choferes['CHOFER'].unique() if str(x).strip() != ""])
                chofer_sel = st.selectbox("🔍 Buscar Perfil Operativo (Chofer):", ["Resumen General"] + lista_choferes, key="sb_chofer")
                
                if chofer_sel == "Resumen General":
                    st.subheader(f"📊 Métricas Globales de Choferes ({mes_seleccionado})")
                    total_choferes = len(lista_choferes)
                    
                    hoy_str = datetime.now().strftime('%d/%m/%Y')
                    inactivos_hoy = 0
                    if 'FECHA' in df_choferes.columns:
                        df_hoy = df_choferes[df_choferes['FECHA'] == hoy_str]
                        if not df_hoy.empty:
                            condicion_inactivo = df_hoy['OBSERVACIÓN'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False) | df_hoy['RUTA'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False)
                            inactivos_hoy = condicion_inactivo.sum()
                    
                    c1, c2, c3 = st.columns(3)
                    with c1: st.metric("Total Choferes Registrados", total_choferes)
                    with c2: st.metric("Choferes Inactivos (Hoy)", inactivos_hoy)
                    with c3: st.metric("Registros en el Periodo", len(df_choferes))
                else:
                    df_ind = df_choferes[df_choferes['CHOFER'] == chofer_sel].copy()
                    
                    unidades_disp = ["Todas"] + sorted([str(x) for x in df_ind['UNIDAD'].unique() if str(x).strip() != ""])
                    zonas_disp = ["Todas"] + sorted([str(x) for x in df_ind['ZONA'].unique() if str(x).strip() != ""])
                    
                    c_f1, c_f2 = st.columns(2)
                    with c_f1:
                        unidad_sel = st.selectbox("🚛 Filtrar por Unidad (Opcional):", unidades_disp, key="u_ch")
                    with c_f2:
                        zona_sel = st.selectbox("📍 Filtrar por Zona (Opcional):", zonas_disp, key="z_ch")
                        
                    if unidad_sel != "Todas": df_ind = df_ind[df_ind['UNIDAD'] == unidad_sel]
                    if zona_sel != "Todas": df_ind = df_ind[df_ind['ZONA'] == zona_sel]
                    
                    texto_filtros = ""
                    if unidad_sel != "Todas": texto_filtros += f" | Und: {unidad_sel}"
                    if zona_sel != "Todas": texto_filtros += f" | Zona: {zona_sel}"

                    df_ind = df_ind.sort_values(by='FECHA_DT', ascending=True)
                    total_dias = len(df_ind)
                    condicion = df_ind['OBSERVACIÓN'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False) | df_ind['RUTA'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False)
                    dias_inactivos = condicion.sum()
                    dias_activos = total_dias - dias_inactivos

                    st.markdown(f"### 👤 Perfil Operativo: {chofer_sel}")
                    c1, c2, c3 = st.columns(3)
                    with c1: st.metric("Días Registrados (Total)", total_dias)
                    with c2: st.metric("✅ Días Trabajados", dias_activos)
                    with c3: st.metric("⚠️ Días Inactivos (Vac/Reposo)", dias_inactivos)

                    st.divider()
                    col_graf, col_datos = st.columns([0.4, 0.6])
                    
                    with col_graf:
                        st.write("**Distribución por Zona Trabajada**")
                        df_activos = df_ind[~condicion]
                        if not df_activos.empty and 'ZONA' in df_activos.columns:
                            zonas_count = df_activos['ZONA'].value_counts().reset_index()
                            zonas_count.columns = ['ZONA', 'Días']
                            fig = px.pie(zonas_count, values='Días', names='ZONA', hole=0.4, color_discrete_sequence=px.colors.sequential.Blues_r)
                            fig.update_traces(textposition='inside', textinfo='percent+label')
                            fig.update_layout(showlegend=False, margin=dict(t=0, b=0, l=0, r=0), height=300)
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("Sin registros de zona para esta selección.")

                    with col_datos:
                        st.write("**Historial Detallado**")
                        columnas_mostrar = ['FECHA', 'DIA', 'UNIDAD', 'RUTA', 'ZONA', 'OBSERVACIÓN']
                        cols_existentes = [c for c in columnas_mostrar if c in df_ind.columns]
                        df_view = df_ind[cols_existentes]
                        
                        tabla_html_chofer = df_view.style.apply(estilo_personal_html, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
                        st.markdown(tabla_html_chofer, unsafe_allow_html=True)
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        pdf_bytes = crear_pdf_operativo(chofer_sel, "Chofer", df_view, total_dias, dias_activos, dias_inactivos, mes_seleccionado, texto_filtros)
                        st.download_button(label=f"📥 Descargar PDF Gerencial", data=pdf_bytes, file_name=f"Perfil_{chofer_sel.replace(' ', '_')}.pdf", mime="application/pdf", use_container_width=True)

        with tab_ayudantes:
            if not df_ayudantes.empty:
                lista_ayudantes = sorted([str(x) for x in df_ayudantes['AYUDANTE'].unique() if str(x).strip() != ""])
                ayu_sel = st.selectbox("🔍 Buscar Perfil Operativo (Ayudante):", ["Resumen General"] + lista_ayudantes, key="sb_ayu")
                
                if ayu_sel == "Resumen General":
                    st.subheader(f"📊 Métricas Globales de Ayudantes ({mes_seleccionado})")
                    total_ayudantes = len(lista_ayudantes)
                    c1, c2, c3 = st.columns(3)
                    with c1: st.metric("Total Ayudantes Registrados", total_ayudantes)
                    with c2: st.metric("Registros en el Periodo", len(df_ayudantes))
                else:
                    df_ind = df_ayudantes[df_ayudantes['AYUDANTE'] == ayu_sel].copy()
                    
                    unidades_disp_a = ["Todas"] + sorted([str(x) for x in df_ind['UNIDAD'].unique() if str(x).strip() != ""])
                    zonas_disp_a = ["Todas"] + sorted([str(x) for x in df_ind['ZONA'].unique() if str(x).strip() != ""])
                    
                    c_f1, c_f2 = st.columns(2)
                    with c_f1:
                        unidad_sel_a = st.selectbox("🚛 Filtrar por Unidad (Opcional):", unidades_disp_a, key="u_ayu")
                    with c_f2:
                        zona_sel_a = st.selectbox("📍 Filtrar por Zona (Opcional):", zonas_disp_a, key="z_ayu")
                        
                    if unidad_sel_a != "Todas": df_ind = df_ind[df_ind['UNIDAD'] == unidad_sel_a]
                    if zona_sel_a != "Todas": df_ind = df_ind[df_ind['ZONA'] == zona_sel_a]
                    
                    texto_filtros_a = ""
                    if unidad_sel_a != "Todas": texto_filtros_a += f" | Und: {unidad_sel_a}"
                    if zona_sel_a != "Todas": texto_filtros_a += f" | Zona: {zona_sel_a}"

                    df_ind = df_ind.sort_values(by='FECHA_DT', ascending=True)
                    total_dias = len(df_ind)
                    condicion = df_ind['OBSERVACIÓN'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False) | df_ind['RUTA'].astype(str).str.contains('VACACIONES|REPOSO|FALTA', case=False, na=False)
                    dias_inactivos = condicion.sum()
                    dias_activos = total_dias - dias_inactivos

                    st.markdown(f"### 👷 Perfil Operativo: {ayu_sel}")
                    c1, c2, c3 = st.columns(3)
                    with c1: st.metric("Días Registrados (Total)", total_dias)
                    with c2: st.metric("✅ Días Trabajados", dias_activos)
                    with c3: st.metric("⚠️ Días Inactivos (Vac/Reposo)", dias_inactivos)

                    st.divider()
                    col_graf, col_datos = st.columns([0.4, 0.6])
                    
                    with col_graf:
                        st.write("**Distribución por Zona Trabajada**")
                        df_activos = df_ind[~condicion]
                        if not df_activos.empty and 'ZONA' in df_activos.columns:
                            zonas_count = df_activos['ZONA'].value_counts().reset_index()
                            zonas_count.columns = ['ZONA', 'Días']
                            fig = px.pie(zonas_count, values='Días', names='ZONA', hole=0.4, color_discrete_sequence=px.colors.sequential.Oranges_r)
                            fig.update_traces(textposition='inside', textinfo='percent+label')
                            fig.update_layout(showlegend=False, margin=dict(t=0, b=0, l=0, r=0), height=300)
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("Sin registros de zona para esta selección.")

                    with col_datos:
                        st.write("**Historial Detallado**")
                        columnas_mostrar = ['FECHA', 'DIA', 'CHOFER', 'UNIDAD', 'RUTA', 'ZONA', 'OBSERVACIÓN']
                        cols_existentes = [c for c in columnas_mostrar if c in df_ind.columns]
                        df_view = df_ind[cols_existentes]
                        
                        tabla_html_ayu = df_view.style.apply(estilo_personal_html, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
                        st.markdown(tabla_html_ayu, unsafe_allow_html=True)
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        pdf_bytes = crear_pdf_operativo(ayu_sel, "Ayudante", df_view, total_dias, dias_activos, dias_inactivos, mes_seleccionado, texto_filtros_a)
                        st.download_button(label=f"📥 Descargar PDF Gerencial", data=pdf_bytes, file_name=f"Perfil_{ayu_sel.replace(' ', '_')}.pdf", mime="application/pdf", use_container_width=True)

    except Exception as e:
        st.error(f"Error cargando los datos de Personal: {e}")

# --- 7.5. MÓDULO 3: TORRE DE CONTROL (DASHBOARD GERENCIAL) ---
@st.fragment(run_every="5m")
def modulo_torre_control():
    col_titulo, col_boton, col_reloj = st.columns([0.60, 0.20, 0.20])
    with col_titulo:
        st.title("📡 MÓDULO: TORRE DE CONTROL DROTACA")
    with col_boton:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Pizarras", use_container_width=True):
            st.rerun() 
            
    with col_reloj:
        if st.session_state.usuario_actual in ["David_Admin", "Jsuarez", "Franluis_pulve"]:
            st.markdown("<br>", unsafe_allow_html=True)
            id_reloj = time.time() 
            codigo_html_reloj = f"""
            <div style="font-family: sans-serif; text-align: center; background-color: #E8F5E9; padding: 5px; border-radius: 8px; border: 2px solid #198754; color: #198754; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                <span style="font-size: 11px; font-weight: bold; text-transform: uppercase;">⏱️ Próxima lectura en:</span><br>
                <span id="reloj_{id_reloj}" style="font-size: 22px; font-weight: 900; letter-spacing: 2px;">05:00</span>
            </div>
            <script>
                let tiempo = 300; 
                const elemento = document.getElementById('reloj_{id_reloj}');
                const intervalo = setInterval(() => {{
                    tiempo--;
                    if (tiempo <= 0) {{
                        tiempo = 0; 
                        elemento.innerText = "Cargando...";
                        elemento.style.fontSize = "16px";
                        clearInterval(intervalo);
                    }} else {{
                        let m = Math.floor(tiempo / 60);
                        let s = tiempo % 60;
                        elemento.innerText = (m < 10 ? '0' : '') + m + ':' + (s < 10 ? '0' : '') + s;
                    }}
                }}, 1000);
            </script>
            """
            components.html(codigo_html_reloj, height=70)
            
    st.markdown("---")

    config_pizarras_todas = {
        "Occidente": {"titulo": "RUTA OCCIDENTE", "responsable": "Jesus Brito"},
        "Centro": {"titulo": "RUTA CENTRO", "responsable": "Jerald Poche"},
        "Oriente": {"titulo": "RUTA ORIENTE", "responsable": "Gabriel Vera"}
    }

    usuario = st.session_state.usuario_actual
    if usuario == "Supervisor_Oriente":
        config_pizarras = {"Oriente": config_pizarras_todas["Oriente"]}
    elif usuario == "Supervisor_Centro":
        config_pizarras = {"Centro": config_pizarras_todas["Centro"]}
    elif usuario == "Supervisor_Occidente":
        config_pizarras = {"Occidente": config_pizarras_todas["Occidente"]}
    else:
        config_pizarras = config_pizarras_todas

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    libro = None
    for intento in range(3):
        try:
            cliente = gspread.authorize(creds)
            libro = cliente.open("Sistema_Flota_2026")
            break 
        except Exception as e:
            if intento < 2:
                time.sleep(3) 
            else:
                st.warning("⚠️ Google Sheets está ocupado sincronizando las pizarras locales. Se recuperará automáticamente en el próximo ciclo.")
                return

    novedades_del_dia = {"Occidente": 0, "Centro": 0, "Oriente": 0}
    fecha_hoy_str = obtener_hora_venezuela().strftime("%d/%m/%Y")
    
    try:
        ws_novedades = libro.worksheet("Novedades_Ruta")
        datos_novedades = ws_novedades.get_all_records()
        if datos_novedades:
            df_n = pd.DataFrame(datos_novedades)
            if 'FECHA' in df_n.columns and 'ZONA' in df_n.columns:
                df_hoy = df_n[df_n['FECHA'].astype(str).str.strip() == fecha_hoy_str]
                for z in df_hoy['ZONA'].dropna():
                    z_str = str(z).strip().upper()
                    if "OCCIDENTE" in z_str: novedades_del_dia["Occidente"] += 1
                    elif "CENTRO" in z_str: novedades_del_dia["Centro"] += 1
                    elif "ORIENTE" in z_str: novedades_del_dia["Oriente"] += 1
    except Exception:
        pass

    datos_por_zona = {}
    global_cubiertos = 0
    global_bultos = 0

    for zona, info_zona in config_pizarras.items():
        try:
            ws_zona = libro.worksheet(f"Pizarra_{zona}")
            datos = ws_zona.get_all_records()
            if datos:
                df = pd.DataFrame(datos)
                
                hora_act = "No registrada"
                if "ULTIMA ACTUALIZACION" in df.columns:
                    hora_act = df["ULTIMA ACTUALIZACION"].iloc[0]
                    df = df.drop(columns=["ULTIMA ACTUALIZACION"])
                
                c_cubiertos = next((c for c in df.columns if 'CUBIERTOS' in str(c).upper()), None)
                c_bultos = next((c for c in df.columns if 'BULTOS' in str(c).upper()), None)
                
                if c_cubiertos: global_cubiertos += pd.to_numeric(df[c_cubiertos], errors='coerce').fillna(0).sum()
                if c_bultos: global_bultos += pd.to_numeric(df[c_bultos], errors='coerce').fillna(0).sum()
                
                datos_por_zona[zona] = {"df": df, "hora": hora_act, "responsable": info_zona["responsable"], "titulo": info_zona["titulo"]}
        except:
            pass

    if usuario in ["David_Admin", "Jsuarez", "Franluis_pulve"]:
        st.markdown("### 🌎 Panorama Operativo Global")
    else:
        zona_texto = list(config_pizarras.keys())[0]
        st.markdown(f"### 🌎 Panorama Operativo: RUTA {zona_texto.upper()}")
        
    col_g1, col_g2, col_g3 = st.columns(3)
    col_g1.metric("🗓️ Jornada Activa", fecha_hoy_str)
    col_g2.metric("✅ Alcance de Clientes", f"{int(global_cubiertos)} Clientes")
    col_g3.metric("📦 Volumen Distribuido", f"{int(global_bultos)} Bultos")
    st.markdown("---")

    for zona, data in datos_por_zona.items():
        st.markdown(f"### 📍 PIZARRA {data['titulo']}")
        col_res, col_hora, _ = st.columns([2, 2, 4])
        with col_res:
            st.markdown(f"👤 **Responsable:** `{data['responsable']}`")
        with col_hora:
            if data['hora'] and str(data['hora']).strip() != "" and str(data['hora']).strip() != "nan":
                st.markdown(f"⏱️ **Actualizado:** `{data['hora']}`")
            else:
                st.markdown(f"⚠️ **Actualizado:** `Pendiente`")

        df_zona = data["df"].copy()
        
        c_cubrir = next((c for c in df_zona.columns if 'A CUBRIR' in str(c).upper()), None)
        c_cubiertos = next((c for c in df_zona.columns if 'CUBIERTOS' in str(c).upper()), None)
        c_pendientes = next((c for c in df_zona.columns if 'PENDIENTES' in str(c).upper()), None)
        c_bultos = next((c for c in df_zona.columns if 'BULTOS' in str(c).upper()), None)
        primer_columna = df_zona.columns[0]
        
        sum_cubrir = pd.to_numeric(df_zona[c_cubrir], errors='coerce').fillna(0).sum() if c_cubrir else 0
        sum_cubiertos = pd.to_numeric(df_zona[c_cubiertos], errors='coerce').fillna(0).sum() if c_cubiertos else 0
        sum_pendientes = pd.to_numeric(df_zona[c_pendientes], errors='coerce').fillna(0).sum() if c_pendientes else 0
        sum_bultos = pd.to_numeric(df_zona[c_bultos], errors='coerce').fillna(0).sum() if c_bultos else 0

        fila_totales = {col: "" for col in df_zona.columns}
        fila_totales[primer_columna] = "TOTALES"
        if c_cubrir: fila_totales[c_cubrir] = int(sum_cubrir)
        if c_cubiertos: fila_totales[c_cubiertos] = int(sum_cubiertos)
        if c_pendientes: fila_totales[c_pendientes] = int(sum_pendientes)
        if c_bultos: fila_totales[c_bultos] = int(sum_bultos)

        df_zona = pd.concat([df_zona, pd.DataFrame([fila_totales])], ignore_index=True)

        def estilo_pizarra_html(row):
            styles = [''] * len(row)
            if row[primer_columna] == 'TOTALES':
                for i, col in enumerate(row.index):
                    base_style_total = 'font-weight: bold; text-align: center; border: 1px solid black; font-size: 20px; '
                    if col == c_bultos:
                        styles[i] = base_style_total + 'background-color: #FACC15; color: black;'
                    else:
                        styles[i] = base_style_total + 'background-color: #1A3B5C; color: white;'
            else:
                for i, col in enumerate(row.index):
                    base_style = 'text-align: center; border: 1px solid black; background-color: white; color: black; '
                    if col in [c_cubrir, c_cubiertos, c_pendientes, c_bultos] and col:
                        base_style += 'font-size: 20px; font-weight: bold; '
                    
                    if col == c_pendientes:
                        valor = pd.to_numeric(row[col], errors='coerce')
                        if valor > 0:
                            styles[i] = base_style + 'background-color: #FFEFEF; color: #DC3545;'
                        elif valor == 0:
                            styles[i] = base_style + 'color: #198754;'
                        else:
                            styles[i] = base_style
                    else:
                        styles[i] = base_style
            return styles
        
        col_tabla, col_grafico = st.columns([0.8, 0.2])
        
        with col_tabla:
            tabla_html = df_zona.style.apply(estilo_pizarra_html, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
            st.markdown(tabla_html, unsafe_allow_html=True)
            
        with col_grafico:
            st.markdown(f"<div style='text-align: center; font-weight: bold; color: #1A3B5C; font-size: 16px;'>EFECTIVIDAD</div>", unsafe_allow_html=True)
            if sum_cubrir > 0:
                df_pie = pd.DataFrame({'Estado': ['Entregado', 'Pendiente'], 'Cantidad': [sum_cubiertos, sum_pendientes]})
                fig = px.pie(df_pie, values='Cantidad', names='Estado', hole=0.4, 
                             color='Estado', color_discrete_map={'Entregado':'#4F81BD', 'Pendiente':'#C0504D'})
                fig.update_traces(textposition='inside', textinfo='percent')
                fig.update_layout(showlegend=False, margin=dict(t=10, b=10, l=10, r=10), height=200)
                st.plotly_chart(fig, use_container_width=True, config={'displayModeBar': False})
            else:
                st.info("Sin datos de clientes para calcular efectividad.")
            
            st.markdown("<br>", unsafe_allow_html=True)
            cantidad_nov = novedades_del_dia.get(zona, 0)
            if cantidad_nov == 0:
                st.markdown(f"<div style='text-align: center; background-color: #E8F5E9; color: #198754; padding: 10px; border-radius: 5px; border: 2px solid #198754; font-weight: bold; font-size: 14px;'>🚨 Novedades en Ruta:<br>Sin Novedad</div>", unsafe_allow_html=True)
            else:
                st.markdown(f"<div style='text-align: center; background-color: #FFEFEF; color: #DC3545; padding: 10px; border-radius: 5px; border: 2px solid #DC3545; font-weight: bold; font-size: 15px;'>🚨 Novedades en Ruta:<br>{cantidad_nov:02d}</div>", unsafe_allow_html=True)
                
        st.markdown("<br><br>", unsafe_allow_html=True)

# --- 7.9. MÓDULO 4: NOVEDADES EN RUTA ---
def modulo_novedades():
    col_titulo, col_boton = st.columns([0.8, 0.2])
    with col_titulo:
        st.title("🚨 Historial de Novedades en Ruta")
    with col_boton:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Bitácora", use_container_width=True):
            st.rerun()
    st.markdown("---")

    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    try:
        cliente = gspread.authorize(creds)
        libro = cliente.open("Sistema_Flota_2026")
        ws_novedades = libro.worksheet("Novedades_Ruta")
        datos = ws_novedades.get_all_records()
    except Exception as e:
        st.warning("⚠️ No se encontró la hoja 'Novedades_Ruta' o está vacía. Asegúrate de crearla con las 10 columnas exactas.")
        return

    if not datos:
        st.info("No hay novedades registradas en el sistema todavía.")
        return

    df_nov = pd.DataFrame(datos)
    
    cols_esperadas = ['FECHA', 'HORA', 'RUTA', 'ZONA', 'PLACA', 'UNIDAD', 'CHOFER', 'AYUDANTE', 'TIPO DE NOVEDAD', 'DESCRIPCIÓN']
    for c in cols_esperadas:
        if c not in df_nov.columns:
            df_nov[c] = ""
            
    df_nov = df_nov[cols_esperadas]

    df_nov['FECHA_DT'] = pd.to_datetime(df_nov['FECHA'], format='%d/%m/%Y', errors='coerce')
    df_nov['MES_NUM'] = df_nov['FECHA_DT'].dt.month
    df_nov['MES_NOMBRE'] = df_nov['MES_NUM'].map(MESES_ESPANOL)
    
    meses_validos = [str(m) for m in df_nov['MES_NOMBRE'].dropna().unique()]
    meses_disp = ["Todos los meses"] + sorted(meses_validos, key=lambda m: list(MESES_ESPANOL.values()).index(m) if m in MESES_ESPANOL.values() else 0)

    st.markdown("#### 🔍 Filtros de Búsqueda Avanzados")
    col_f1, col_f2, col_f3, col_f4 = st.columns(4)
    
    rutas_disp = ["Todas"] + sorted([str(x) for x in df_nov['RUTA'].unique() if str(x).strip() != ""])
    choferes_disp = ["Todos"] + sorted([str(x) for x in df_nov['CHOFER'].unique() if str(x).strip() != ""])
    ayudantes_disp = ["Todos"] + sorted([str(x) for x in df_nov['AYUDANTE'].unique() if str(x).strip() != ""])
    
    with col_f1: f_mes = st.selectbox("📅 Filtrar por Mes:", meses_disp)
    with col_f2: f_ruta = st.selectbox("📍 Filtrar por Ruta:", rutas_disp)
    with col_f3: f_chofer = st.selectbox("👤 Filtrar por Chofer:", choferes_disp)
    with col_f4: f_ayudante = st.selectbox("👷 Filtrar por Ayudante:", ayudantes_disp)
    
    f_busqueda = st.text_input("🔎 Búsqueda libre (Placa, Zona, Unidad o Palabra clave):")

    df_filtrado = df_nov.copy()
    filtros_usados = []
    
    if f_mes != "Todos los meses":
        df_filtrado = df_filtrado[df_filtrado['MES_NOMBRE'] == f_mes]
        filtros_usados.append(f"Mes: {f_mes}")
    if f_ruta != "Todas":
        df_filtrado = df_filtrado[df_filtrado['RUTA'] == f_ruta]
        filtros_usados.append(f"Ruta: {f_ruta}")
    if f_chofer != "Todos":
        df_filtrado = df_filtrado[df_filtrado['CHOFER'] == f_chofer]
        filtros_usados.append(f"Chofer: {f_chofer}")
    if f_ayudante != "Todos":
        df_filtrado = df_filtrado[df_filtrado['AYUDANTE'] == f_ayudante]
        filtros_usados.append(f"Ayudante: {f_ayudante}")
    if f_busqueda:
        mask = df_filtrado.astype(str).apply(lambda x: x.str.contains(f_busqueda, case=False)).any(axis=1)
        df_filtrado = df_filtrado[mask]
        filtros_usados.append(f"Búsqueda: '{f_busqueda}'")
        
    texto_filtros = " | ".join(filtros_usados) if filtros_usados else "Ninguno (Mostrando todo)"

    df_filtrado = df_filtrado.drop(columns=['FECHA_DT', 'MES_NUM', 'MES_NOMBRE'])

    st.markdown("<br>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns([1, 2, 1])
    
    c1.metric("Total de Novedades (Filtradas)", len(df_filtrado))
    
    with c3:
        st.markdown("<br>", unsafe_allow_html=True)
        if not df_filtrado.empty:
            pdf_novedades = crear_pdf_novedades(df_filtrado, texto_filtros)
            st.download_button(
                label="📄 Descargar Bitácora en PDF",
                data=pdf_novedades,
                file_name=f"Bitacora_Novedades_{obtener_hora_venezuela().strftime('%d%m%Y')}.pdf",
                mime="application/pdf",
                use_container_width=True,
                type="primary"
            )

    st.markdown("---")
    st.markdown("### 📋 Historial Inmutable")

    if not df_filtrado.empty:
        def estilo_novedades(row):
            styles = ['border: 1px solid black; text-align: center; font-size: 14px; color: black; background-color: white;'] * len(row)
            for i, col in enumerate(row.index):
                if col == 'DESCRIPCIÓN':
                    styles[i] = 'border: 1px solid black; text-align: left; font-size: 14px; padding: 10px; font-style: italic; background-color: #F8F9FA; color: black;' 
            return styles
        
        tabla_html = df_filtrado.style.apply(estilo_novedades, axis=1).set_table_styles(estilos_html_genericos).hide(axis="index").to_html()
        st.markdown(tabla_html, unsafe_allow_html=True)
    else:
        st.info("No hay resultados para los filtros seleccionados.")

# --- 8. CONTROL DE FLUJO Y NAVEGACIÓN ---
if not st.session_state.autenticado:
    pantalla_login()
else:
    st.markdown("""
    <style>
    [data-testid="stApp"] { background: #F0F4F8 !important; color: #31333F !important; }
    [data-testid="stHeader"] { background-color: transparent !important; }
    .stAppDeployButton {display:none;}
    </style>
    """, unsafe_allow_html=True)

    with st.sidebar:
        st.markdown(f"### 👤 Usuario:\n**{st.session_state.nombre_real}**")
        st.divider()
        st.markdown("### 🗂️ Módulos del Sistema")
        menu_seleccionado = st.radio("", ["🚛 Control de Flota", "👥 Rotación de Personal", "🗼 Torre de Control", "🚨 Novedades en Ruta", "🛒 Requisiciones y Compras"])
        
        if st.session_state.usuario_actual == "David_Admin":
            st.divider()
            st.markdown("### 🤖 Control de Bots (Pizarras)")
            try:
                scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
                try:
                    credenciales_dict = dict(st.secrets["gcp_service_account"])
                    creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
                except:
                    creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
                cliente = gspread.authorize(creds)
                libro_config = cliente.open("Sistema_Flota_2026")
                ws_config = libro_config.worksheet("Configuracion")
                estado_actual = str(ws_config.acell('A1').value).strip().upper()
                
                bot_activado = st.toggle("Encender Sincronización Automática", value=(estado_actual == "ENCENDIDO"))
                nuevo_estado = "ENCENDIDO" if bot_activado else "APAGADO"
                
                if nuevo_estado != estado_actual:
                    ws_config.update_acell('A1', nuevo_estado)
                    st.toast(f"Estado de los Bots cambiado a: {nuevo_estado}")
                    st.success(f"Bots en {nuevo_estado}")
            except Exception as e:
                st.caption("⚠️ No se pudo cargar el estado de los bots. Verifica la hoja 'Configuracion'.")
                
        st.divider()
        if st.button("🚪 Cerrar Sesión", use_container_width=True, type="primary"):
            st.session_state.autenticado = False
            st.session_state.usuario_actual = ""
            st.session_state.nombre_real = ""
            st.rerun()

    if menu_seleccionado == "🚛 Control de Flota":
        modulo_flota()
    elif menu_seleccionado == "👥 Rotación de Personal":
        modulo_personal()
    elif menu_seleccionado == "🗼 Torre de Control":
        modulo_torre_control()
    elif menu_seleccionado == "🚨 Novedades en Ruta":
        modulo_novedades()
    elif menu_seleccionado == "🛒 Requisiciones y Compras":
        renderizar_modulo_compras()
