import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime, timedelta
import os
import tempfile
import time
from fpdf import FPDF
import openpyxl
import io

# --- FUNCIONES AUXILIARES INDEPENDIENTES ---
def obtener_hora_venezuela():
    return datetime.utcnow() - timedelta(hours=4)

def limpiar_texto_pdf(texto):
    if pd.isna(texto): return ""
    texto = str(texto)
    reemplazos = {
        "Ñ":"N", "ñ":"n", "á":"a", "é":"e", "í":"i", "ó":"o", "ú":"u",
        "Á":"A", "É":"E", "Í":"I", "Ó":"O", "Ú":"U",
        '\u2013':'-', '\u2014':'-', '\u2018':"'", '\u2019':"'", '\u201c':'"', '\u201d':'"', '\u2026':'...'
    }
    for k, v in reemplazos.items():
        texto = texto.replace(k, v)
    return texto.encode('latin-1', 'ignore').decode('latin-1').strip()

def calcular_dias_sla(row):
    try:
        f_sol_str = str(row.get('FECHA_SOLICITUD', '')).strip()
        if not f_sol_str: return ""
        f_sol = datetime.strptime(f_sol_str, "%d/%m/%Y")
        
        estado = str(row.get('ESTADO', '')).strip().upper()
        
        if 'COMPRADO' in estado:
            f_com_str = str(row.get('FECHA_COMPRA', '')).split(' - ')[0].strip()
            if f_com_str:
                f_com = datetime.strptime(f_com_str, "%d/%m/%Y")
                dias = (f_com - f_sol).days
                return f"✅ {dias} días"
            return "✅ -"
        else:
            hoy = obtener_hora_venezuela().replace(hour=0, minute=0, second=0, microsecond=0)
            dias = (hoy - f_sol).days
            if dias > 5:
                return f"⚠️ {dias} días"
            elif dias < 0:
                return f"⏳ 0 días"
            else:
                return f"⏳ {dias} días"
    except:
        return "-"

def extraer_mes_texto(fecha_str):
    meses = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio', 
             7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
    try:
        dt = datetime.strptime(str(fecha_str).strip(), "%d/%m/%Y")
        return f"{meses[dt.month]} {dt.year}"
    except:
        return "Desconocido"

# --- MEMORIA CACHÉ ---
@st.cache_data(ttl=600, show_spinner=False)
def cargar_datos_compras():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    cliente = gspread.authorize(creds)
    libro = cliente.open("Sistema_Flota_2026")
    
    try:
        ws_req = libro.worksheet("Requisiciones_Taller")
        df_req = pd.DataFrame(ws_req.get_all_records())
    except:
        df_req = pd.DataFrame(columns=['ID_REQ', 'FECHA_SOLICITUD', 'HORA_SOLICITUD', 'SEMANA', 'ZONA', 'UNIDAD', 'SOLICITANTE', 'CANTIDAD', 'DESCRIPCION_ITEM', 'NOTA_SOLICITUD', 'ESTADO', 'FECHA_COMPRA', 'COMPRADOR', 'NOTA_COMPRA', 'VALOR', 'FACTURA', 'TIPO_REQ'])
        
    cols_necesarias = ['VALOR', 'FACTURA', 'TIPO_REQ']
    for col in cols_necesarias:
        if col not in df_req.columns:
            df_req[col] = ""

    try:
        ws_maestro = libro.worksheet("Maestro_Flota")
        df_maestro = pd.DataFrame(ws_maestro.get_all_records())
        placas_lista = df_maestro['Placa'].dropna().astype(str).unique().tolist()
        placas_lista = [p.strip().upper() for p in placas_lista if p.strip()]
        placas_lista.sort()
    except:
        placas_lista = []
        
    return df_req, placas_lista

# --- FUNCIÓN PARA GUARDAR DATOS ---
def guardar_accion_bd(accion, detalle, nueva_fila=None, actualizar_datos=None):
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    try:
        credenciales_dict = dict(st.secrets["gcp_service_account"])
        creds = ServiceAccountCredentials.from_json_keyfile_dict(credenciales_dict, scope)
    except:
        creds = ServiceAccountCredentials.from_json_keyfile_name("credenciales.json", scope)
        
    cliente = gspread.authorize(creds)
    libro = cliente.open("Sistema_Flota_2026")
    ws_req = libro.worksheet("Requisiciones_Taller")
    ws_aud = libro.worksheet("Auditoria_Compras")
    
    if nueva_fila:
        ws_req.append_row(nueva_fila)
        
    if actualizar_datos:
        id_sel, estado_bd, ahora_str, usuario, nota_bd, valor, factura = actualizar_datos
        col_ids = ws_req.col_values(1)
        fila_actualizar = col_ids.index(id_sel) + 1 
        ws_req.update_cell(fila_actualizar, 11, estado_bd)
        ws_req.update_cell(fila_actualizar, 12, ahora_str)
        ws_req.update_cell(fila_actualizar, 13, usuario)
        ws_req.update_cell(fila_actualizar, 14, nota_bd)
        ws_req.update_cell(fila_actualizar, 15, valor)
        ws_req.update_cell(fila_actualizar, 16, factura)
        
    ahora_aud = obtener_hora_venezuela()
    nombre_auditoria = st.session_state.get('nombre_real', st.session_state.usuario_actual)
    ws_aud.append_row([ahora_aud.strftime("%d/%m/%Y"), ahora_aud.strftime("%I:%M %p"), nombre_auditoria, accion, detalle])

# --- MOTORES GENERADORES DE PDF ---
def crear_pdf_historial(df_datos, texto_filtros):
    class PDFHistorial(FPDF):
        def header(self):
            if os.path.exists("encabezado.png"):
                self.image("encabezado.png", x=10, y=8, w=277)
                self.set_y(46) 
            else:
                self.set_y(15)
                
            self.set_font('Arial', 'B', 15)
            self.set_fill_color(26, 59, 92)
            self.set_text_color(255, 255, 255)
            self.cell(0, 12, ' DROTACA - HISTORIAL DE REQUISICIONES', 0, 1, 'C', 1)
            self.ln(4)

    pdf = PDFHistorial('L', 'mm', 'A4') 
    pdf.add_page()
    pdf.set_font('Arial', 'B', 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, f"Filtros: {limpiar_texto_pdf(texto_filtros)}", 0, 1)
    pdf.cell(0, 5, f"Total Registros: {len(df_datos)}", 0, 1)
    pdf.ln(3)

    titulos_pdf = ['ID_REQ', 'TIPO', 'F. SOLIC', 'F. COMPRA', 'UNIDAD', 'SOLICITA', 'CANT.', 'DESCRIPCION', 'VALOR', 'FACTURA', 'ESTADO']
    anchos = [25, 25, 18, 18, 15, 20, 22, 62, 15, 24, 33] 
    
    pdf.set_font('Arial', 'B', 8)
    pdf.set_fill_color(230, 230, 230)
    pdf.set_text_color(0, 0, 0)
    for i, col in enumerate(titulos_pdf):
        pdf.cell(anchos[i], 8, limpiar_texto_pdf(col), 1, 0, 'C', 1)
    pdf.ln()

    columnas_df = ['ID_REQ', 'TIPO_REQ', 'FECHA_SOLICITUD', 'FECHA_COMPRA', 'UNIDAD', 'SOLICITANTE', 'CANTIDAD', 'DESCRIPCION_ITEM', 'VALOR', 'FACTURA', 'ESTADO']
    
    for index, row in df_datos.iterrows():
        pdf.set_font('Arial', 'B', 7)
        pdf.set_text_color(0, 0, 0)
        estado_str = str(row.get('ESTADO', '')).strip().upper()
        if 'PENDIENTE' in estado_str: pdf.set_text_color(220, 53, 69)
        elif 'COMPRADO' in estado_str: pdf.set_text_color(25, 135, 84)
        else: pdf.set_text_color(0, 0, 0)

        for i, col in enumerate(columnas_df):
            if col == 'FECHA_COMPRA':
                val = str(row.get(col, '')).split(' - ')[0].strip()
            else:
                val = str(row.get(col, ''))
                
            texto = limpiar_texto_pdf(val)
            alineacion = 'L' if col in ['DESCRIPCION_ITEM', 'SOLICITANTE', 'TIPO_REQ'] else 'C'
            texto_cortado = texto[:int(anchos[i] * 0.75)]
            pdf.cell(anchos[i], 6, texto_cortado, 1, 0, alineacion)
        pdf.ln()
        
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    os.remove(tmp.name)
    return data

def crear_pdf_planilla_oficial(df_seleccionados, notas_combinadas):
    class PDFPlanilla(FPDF):
        def header(self):
            if os.path.exists("encabezado.png"):
                self.image("encabezado.png", x=10, y=8, w=196)
                self.set_y(38)
            else:
                self.set_y(15)

            self.set_font('Arial', 'B', 16)
            self.cell(0, 8, 'PLANILLA DE SOLICITUD DE COMPRA', 0, 1, 'C')
            self.set_font('Arial', 'I', 10)
            self.cell(0, 5, 'Repuestos, Consumibles, Herramientas para uso Mecanico', 0, 1, 'C')
            self.ln(6)

    pdf = PDFPlanilla('P', 'mm', 'Letter')
    pdf.add_page()
    
    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(220, 220, 220)
    pdf.cell(20, 8, 'N DE ITEM', 1, 0, 'C', 1)
    pdf.cell(45, 8, 'CANTIDAD SOLICITADA', 1, 0, 'C', 1)
    pdf.cell(130, 8, 'DESCRIPCION', 1, 1, 'C', 1)
    
    pdf.set_font('Arial', '', 9)
    for i in range(13):
        if i < len(df_seleccionados):
            row = df_seleccionados.iloc[i]
            cant = str(row.get('CANTIDAD', ''))
            desc = str(row.get('DESCRIPCION_ITEM', ''))
        else:
            cant = ""
            desc = ""
            
        pdf.cell(20, 7, str(i+1), 1, 0, 'C')
        pdf.cell(45, 7, limpiar_texto_pdf(cant), 1, 0, 'C')
        pdf.cell(130, 7, limpiar_texto_pdf(desc), 1, 1, 'L')
        
    pdf.ln(4)
    
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(195, 5, 'NOTA:', 'L T R', 1, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.multi_cell(195, 12, limpiar_texto_pdf(notas_combinadas), 'L B R', 'L')
    
    pdf.ln(8)
    x_start = pdf.get_x()
    y_start = pdf.get_y()
    
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(85, 6, 'SOLICITADO', 1, 1, 'C', 1)
    pdf.set_font('Arial', '', 9)
    pdf.cell(85, 6, 'Nombre: Franluis Pulve', 'L R', 1, 'L')
    pdf.cell(85, 6, 'Analista de Flota y Logistica', 'L R', 1, 'L')
    pdf.cell(85, 6, 'Firma: .................................................', 'L R', 1, 'L')
    pdf.cell(85, 6, 'Fecha: .................................................', 'L B R', 1, 'L')
    
    pdf.set_xy(x_start + 110, y_start)
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(85, 6, 'RECIBIDO', 1, 1, 'C', 1)
    pdf.set_x(x_start + 110)
    pdf.set_font('Arial', '', 9)
    pdf.cell(85, 6, 'Nombre: Javier Hidalgo', 'L R', 1, 'L')
    pdf.set_x(x_start + 110)
    pdf.cell(85, 6, 'Supervisor de Flota y Logistica', 'L R', 1, 'L')
    pdf.set_x(x_start + 110)
    pdf.cell(85, 6, 'Firma: .................................................', 'L R', 1, 'L')
    pdf.set_x(x_start + 110)
    pdf.cell(85, 6, 'Fecha: .................................................', 'L B R', 1, 'L')
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
        pdf.output(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    os.remove(tmp.name)
    return data

# --- FUNCIÓN PRINCIPAL DEL MÓDULO ---
def renderizar_modulo_compras():
    st.markdown("""
    <style>
        div[data-baseweb="input"] > div, 
        div[data-baseweb="textarea"] > div,
        div[data-baseweb="select"] > div {
            background-color: #1A3B5C !important;
            border: 2px solid black !important; 
            border-radius: 5px !important;
        }
        div[data-baseweb="input"] input, 
        div[data-baseweb="textarea"] textarea,
        div[data-baseweb="select"] * {
            color: #FFFFFF !important;
            -webkit-text-fill-color: #FFFFFF !important;
            font-weight: bold !important;
        }
        div[data-baseweb="input"] input::placeholder, 
        div[data-baseweb="textarea"] textarea::placeholder {
            color: #E0E0E0 !important;
            -webkit-text-fill-color: #E0E0E0 !important;
            opacity: 1 !important;
        }
        ul[data-baseweb="menu"] {
            background-color: #1A3B5C !important;
            border: 2px solid black !important;
        }
        li[role="option"] {
            color: #FFFFFF !important;
            background-color: #1A3B5C !important;
            font-weight: bold !important;
        }
        li[role="option"]:hover {
            background-color: #112840 !important; 
        }
        div[data-baseweb="input"] > div:focus-within, 
        div[data-baseweb="textarea"] > div:focus-within,
        div[data-baseweb="select"] > div:focus-within {
            border: 2px solid black !important; 
            box-shadow: none !important; 
        }
    </style>
    """, unsafe_allow_html=True)

    col_titulo, col_boton = st.columns([0.8, 0.2])
    with col_titulo:
        st.title("🛒 Gestión de Repuestos y Compras")
    with col_boton:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🔄 Actualizar Datos", use_container_width=True):
            cargar_datos_compras.clear()
            st.rerun()
    st.markdown("---")

    try:
        df_req, placas_lista = cargar_datos_compras()
        placas_lista.append("PARTICULAR") 
    except Exception as e:
        st.error("Error conectando a la base de datos de compras.")
        return
        
    if not df_req.empty:
        df_req['TIEMPO_SLA'] = df_req.apply(calcular_dias_sla, axis=1)
        df_req['MES'] = df_req['FECHA_SOLICITUD'].apply(extraer_mes_texto)
    else:
        df_req['TIEMPO_SLA'] = []
        df_req['MES'] = []

    tab1, tab2, tab3, tab4 = st.tabs(["📝 Generar Solicitud", "🧾 Registro de Compras y Facturación", "📊 Historial y Auditoría", "🖨️ Generador de Planillas"])

    # --- PESTAÑA 1: SOLICITUD ---
    with tab1:
        st.markdown("### 📝 Formulario de Requisición de Taller")
        st.info("💡 Selecciona el tipo de requisición y rellena los datos a continuación.")
        
        lista_zonas = ["Oriente", "Centro", "Occidente", "Encomiendas", "Transbordo", "Transporte", "Particular"]
        
        # =================================================================================
        # MEJORA: EL SELECTOR DE TIPO ESTÁ AFUERA DEL FORMULARIO PARA QUE SEA DINÁMICO
        # =================================================================================
        st.markdown("#### 📌 1. Tipo de Requisición")
        tipo_req = st.selectbox("Clasificación de la Compra:", ["Emergencia / Unidad (Gasto Directo)", "Reposición de Stock (Almacén)"])
        
        with st.form("form_nueva_solicitud", clear_on_submit=True):
            st.markdown("#### ⚙️ 2. Datos de Operación")
            
            if tipo_req == "Reposición de Stock (Almacén)":
                st.info("📦 Compra para Stock: La Zona y Unidad se asignan automáticamente al Almacén.")
                c1, c2, c3 = st.columns(3)
                with c1: 
                    st.text_input("Zona Operativa:", value="ALMACEN DROTACA 2.0", disabled=True)
                    zona_sol = "ALMACEN DROTACA 2.0"
                    pertenece_zona = ""
                with c2: 
                    st.text_input("Unidad / Placa:", value="ALMACEN DROTACA 2.0", disabled=True)
                    unidad_sol = "ALMACEN DROTACA 2.0"
                    placa_particular = ""
                with c3: 
                    cant_sol = st.text_input("Cantidad (Ej. 5 Unidades o 3 Kits)")
            else:
                c1, c2, c3 = st.columns(3)
                with c1: 
                    zona_sol = st.selectbox("Zona Operativa:", lista_zonas)
                    pertenece_zona = st.text_input("👤 Si la zona es Particular, indique a quién pertenece:")
                with c2: 
                    unidad_sol = st.selectbox("Unidad / Placa:", ["Seleccione..."] + placas_lista)
                    placa_particular = st.text_input("🚗 Si eligió placa PARTICULAR, escríbala aquí:")
                with c3: 
                    cant_sol = st.text_input("Cantidad (Ej. 5 Unidades o 3 Kits)")
                
            st.markdown("#### 🔧 3. Detalles del Repuesto")
            desc_sol = st.text_input("Descripción del Repuesto / Consumible")
            nota_sol = st.text_area("Nota u Observación (Opcional)")
            
            submit_sol = st.form_submit_button("Generar Solicitud", type="primary", use_container_width=True)
            
            if submit_sol:
                if not desc_sol or not cant_sol or (tipo_req == "Emergencia / Unidad (Gasto Directo)" and unidad_sol == "Seleccione..."):
                    st.error("⚠️ Debes llenar la Unidad, Cantidad y Descripción para procesar la solicitud.")
                elif tipo_req == "Emergencia / Unidad (Gasto Directo)" and zona_sol == "Particular" and not pertenece_zona:
                    st.error("⚠️ Seleccionaste Zona Particular, debes indicar a quién pertenece.")
                elif tipo_req == "Emergencia / Unidad (Gasto Directo)" and unidad_sol == "PARTICULAR" and not placa_particular:
                    st.error("⚠️ Seleccionaste Placa Particular, debes escribir la placa abajo.")
                else:
                    with st.spinner("Enviando a Google Sheets..."):
                        ahora = obtener_hora_venezuela()
                        id_unico = f"REQ-{ahora.strftime('%y%m%d%H%M%S')}" 
                        semana = ahora.isocalendar()[1]
                        
                        if tipo_req == "Reposición de Stock (Almacén)":
                            zona_final = "ALMACEN DROTACA 2.0"
                            unidad_final = "ALMACEN DROTACA 2.0"
                        else:
                            zona_final = f"Particular ({pertenece_zona})" if zona_sol == "Particular" else zona_sol
                            unidad_final = placa_particular.upper() if unidad_sol == "PARTICULAR" else unidad_sol
                        
                        nombre_solicitante = st.session_state.get('nombre_real', st.session_state.usuario_actual)
                        
                        nueva_fila = [
                            id_unico, ahora.strftime("%d/%m/%Y"), ahora.strftime("%I:%M %p"), f"Semana {semana}",
                            zona_final.upper(), unidad_final, nombre_solicitante, cant_sol.upper(), 
                            desc_sol.upper(), nota_sol, "Pendiente", "", "", "", "", "", tipo_req
                        ]
                        try:
                            guardar_accion_bd("CREACIÓN", f"Creó solicitud {id_unico} para {unidad_final}", nueva_fila=nueva_fila)
                            cargar_datos_compras.clear() 
                            st.success(f"✅ ¡Solicitud enviada con éxito! Ticket Generado: {id_unico}")
                            time.sleep(2)
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error al guardar: {e}")

    # --- PESTAÑA 2: CONFIRMAR COMPRA ---
    with tab2:
        st.markdown("### 🧾 Registro de Compras y Facturación")
        pendientes = df_req[df_req['ESTADO'].astype(str).str.upper() == 'PENDIENTE']
        
        cantidad_pendientes = len(pendientes)
        
        if cantidad_pendientes == 0:
            st.markdown("<div style='background-color: #E8F5E9; color: #198754; padding: 12px; border-radius: 5px; border: 2px solid #198754; font-weight: bold; text-align: center; margin-bottom: 15px; font-size: 16px;'>✅ ¡Al Día! 0 Solicitudes Pendientes</div>", unsafe_allow_html=True)
        else:
            st.markdown(f"<div style='background-color: #FFF3CD; color: #856404; padding: 12px; border-radius: 5px; border: 2px solid #FFE69C; font-weight: bold; text-align: center; margin-bottom: 15px; font-size: 16px;'>⏳ Tienes {cantidad_pendientes} Solicitudes Pendientes por Procesar</div>", unsafe_allow_html=True)
            
            opciones_pendientes = pendientes['ID_REQ'] + " | " + pendientes['UNIDAD'] + " - " + pendientes['DESCRIPCION_ITEM']
            req_seleccionada = st.selectbox("Seleccione la solicitud a confirmar:", opciones_pendientes.tolist())
            
            if req_seleccionada:
                id_sel = req_seleccionada.split(" | ")[0]
                datos_fila = pendientes[pendientes['ID_REQ'] == id_sel].iloc[0]
                
                tiempo_retraso = str(datos_fila.get('TIEMPO_SLA', ''))
                alerta_retraso = f" — **Lleva {tiempo_retraso} en espera.**" if tiempo_retraso else ""
                tipo_mostrado = str(datos_fila.get('TIPO_REQ', 'No definido'))
                
                st.info(f"📌 **Tipo:** {tipo_mostrado}\n\n👤 **Solicitado por:** {datos_fila['SOLICITANTE']} el {datos_fila['FECHA_SOLICITUD']}{alerta_retraso}\n\n📝 **Nota Original:** {datos_fila['NOTA_SOLICITUD']}")
                st.warning(f"📦 **A comprar:** {datos_fila['CANTIDAD']} x {datos_fila['DESCRIPCION_ITEM']}")
                
                tipo_compra = st.radio("Estado de la Recepción / Compra:", 
                                       ["Completa (Se compró exactamente lo solicitado)", "Parcial / Incompleta (Se compró menos o diferente)"])
                
                st.markdown("#### 💳 Detalles del Pago y Documento")
                
                col_moneda, col_doc = st.columns(2)
                with col_moneda:
                    tipo_moneda = st.radio("Moneda de Pago:", ["Dólares ($)", "Bolívares (Bs)"], horizontal=True)
                with col_doc:
                    tipo_documento = st.radio("Tipo de Documento:", ["Factura Fiscal", "Nota de Entrega"], horizontal=True)

                col_v, col_f = st.columns(2)
                with col_v:
                    valor_input = st.text_input("💰 Monto Total (Solo números. Ej: 55)")
                with col_f:
                    if tipo_documento == "Nota de Entrega":
                        factura_input = st.text_input("🧾 Número de Documento (Opcional si no tiene)")
                    else:
                        factura_input = st.text_input("🧾 Número de Documento (Ej: 855233552)")
                
                val_limpio = valor_input.replace('$', '').replace('Bs', '').replace('bs', '').replace('BS', '').strip()
                simbolo = "$" if "Dólares" in tipo_moneda else " Bs"
                valor_formateado = f"{val_limpio}{simbolo}" if val_limpio else ""
                
                fac_limpia = factura_input.replace('#', '').replace('NE-', '').strip()
                
                if fac_limpia:
                    prefijo = "#" if "Factura" in tipo_documento else "NE-"
                    factura_formateada = f"{prefijo}{fac_limpia}"
                    texto_doc_ws = f"{'Factura Fiscal' if 'Factura' in tipo_documento else 'Nota de Entrega'} {factura_formateada}"
                else:
                    if "Nota de Entrega" in tipo_documento:
                        factura_formateada = "Nota de Entrega"
                        texto_doc_ws = "Nota de Entrega" 
                    else:
                        factura_formateada = ""
                        texto_doc_ws = "No especificado"
                
                col_p1, col_p2 = st.columns(2)
                cant_real = ""
                motivo_parcial = ""
                
                with col_p1:
                    if "Parcial" in tipo_compra: cant_real = st.text_input("⚠️ ¿Qué cantidad compraste realmente? (Ej. 3 unidades)")
                with col_p2:
                    if "Parcial" in tipo_compra: motivo_parcial = st.text_input("⚠️ Motivo de la falta (Ej. No había más en stock)")
                        
                nota_compra = st.text_area("Observación General (Opcional. Ej: Comprado en Repuestos Juan)")
                
                nombre_comprador = st.session_state.get('nombre_real', st.session_state.usuario_actual)
                
                st.markdown("---")
                st.markdown("📲 **TEXTO PARA COPIAR A WHATSAPP:**")
                
                estado_texto = "COMPRADO PARCIALMENTE ⚠️" if "Parcial" in tipo_compra else "COMPRADO ✅"
                cant_mostrar = cant_real if "Parcial" in tipo_compra else datos_fila['CANTIDAD']
                nota_mostrar = f"Motivo: {motivo_parcial} | {nota_compra}" if "Parcial" in tipo_compra else nota_compra
                
                texto_whatsapp = f"""*CONFIRMACIÓN DE COMPRA DROTACA* 🛒
*Ticket:* {id_sel}
*Tipo:* {tipo_mostrado}
*Unidad:* {datos_fila['UNIDAD']}
*Solicitado:* {datos_fila['CANTIDAD']} x {datos_fila['DESCRIPCION_ITEM']}
*Comprado:* {cant_mostrar}
*Valor:* {valor_formateado if valor_formateado else 'No especificado'}
*Documento:* {texto_doc_ws}
*Estatus:* {estado_texto}
*Comprador:* {nombre_comprador}
*Nota:* {nota_mostrar}"""

                st.code(texto_whatsapp, language="text")
                st.markdown("---")
                
                if st.button("✅ Finalizar y Guardar Compra", use_container_width=True):
                    if "Parcial" in tipo_compra and (not cant_real or not motivo_parcial):
                        st.error("⚠️ Para compras parciales debes indicar la cantidad real y el motivo.")
                    else:
                        with st.spinner("Sincronizando con Google Sheets..."):
                            try:
                                ahora_c = obtener_hora_venezuela()
                                estado_bd = "Comprado (Parcial)" if "Parcial" in tipo_compra else "Comprado"
                                nota_bd = f"[PARCIAL] Compró: {cant_real}. Motivo: {motivo_parcial} | Nota: {nota_compra}" if "Parcial" in tipo_compra else nota_compra
                                
                                data_update = (id_sel, estado_bd, ahora_c.strftime("%d/%m/%Y - %I:%M %p"), nombre_comprador, nota_bd, valor_formateado, factura_formateada)
                                guardar_accion_bd("CONFIRMACIÓN", f"Confirmó compra {estado_bd} del ticket {id_sel}", actualizar_datos=data_update)
                                
                                cargar_datos_compras.clear()
                                st.success("✅ Compra registrada correctamente.")
                                time.sleep(2)
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error al actualizar: {e}")

    # --- PESTAÑA 3: HISTORIAL Y AUDITORÍA ---
    with tab3:
        st.markdown("### 📊 Auditoría y Reportes de SLA")
        
        c1, c2, c3, c4 = st.columns(4)
        total_req = len(df_req)
        comprados = len(df_req[df_req['ESTADO'].astype(str).str.contains('COMPRADO', case=False, na=False)])
        pend = total_req - comprados
        efectividad = (comprados / total_req * 100) if total_req > 0 else 0
        criticas = len(df_req[(df_req['ESTADO'].astype(str).str.upper() == 'PENDIENTE') & (df_req['TIEMPO_SLA'].astype(str).str.contains('⚠️'))])
        
        c1.metric("Total Solicitudes", total_req)
        c2.metric("Pendientes", pend)
        c3.metric("Alertas (>5 días)", criticas)
        c4.metric("Efectividad", f"{efectividad:.1f}%")
        
        st.divider()
        st.markdown("#### 🔍 Explorador Maestro Multiltro")
        
        lista_zonas_filtro = ["Todas"] + sorted([z for z in df_req['ZONA'].astype(str).unique().tolist() if z.strip()])
        lista_meses = ["Todos"] + sorted([m for m in df_req['MES'].unique().tolist() if m != "Desconocido"], reverse=True)
        lista_semanas = ["Todas"] + sorted([s for s in df_req['SEMANA'].astype(str).unique().tolist() if s.strip()], reverse=True)
        lista_tipos = ["Todos", "Emergencia / Unidad (Gasto Directo)", "Reposición de Stock (Almacén)"]
        
        cf1, cf2, cf3 = st.columns(3)
        with cf1: f_tipo = st.selectbox("Filtrar por Tipo:", lista_tipos, key="f_tipo_compras")
        with cf2: f_estado = st.selectbox("Filtrar por Estado:", ["Todos", "Pendiente", "Comprado", "Comprado (Parcial)"], key="f_est_compras")
        with cf3: f_zona = st.selectbox("Filtrar por Zona:", lista_zonas_filtro, key="f_zona_compras")
        
        cf4, cf5 = st.columns(2)
        with cf4: f_mes = st.selectbox("Filtrar por Mes:", lista_meses, key="f_mes_compras")
        with cf5: f_semana = st.selectbox("Filtrar por Semana:", lista_semanas, key="f_sem_compras")
        
        f_busqueda_compras = st.text_input("🔎 Búsqueda libre (Buscar Fecha Exacta, Placa, Repuesto o Factura):")

        df_mostrar = df_req.copy()
        filtros_usados_c = []
        
        if f_tipo != "Todos": 
            df_mostrar = df_mostrar[df_mostrar['TIPO_REQ'].astype(str).str.contains(f_tipo[:10], case=False, na=False)]
            filtros_usados_c.append(f"Tipo: {f_tipo[:15]}...")
        if f_estado != "Todos": 
            df_mostrar = df_mostrar[df_mostrar['ESTADO'].astype(str).str.upper() == f_estado.upper()]
            filtros_usados_c.append(f"Estado: {f_estado}")
        if f_zona != "Todas": 
            df_mostrar = df_mostrar[df_mostrar['ZONA'].astype(str).str.contains(f_zona.upper(), case=False, na=False)]
            filtros_usados_c.append(f"Zona: {f_zona}")
        if f_mes != "Todos": 
            df_mostrar = df_mostrar[df_mostrar['MES'] == f_mes]
            filtros_usados_c.append(f"Mes: {f_mes}")
        if f_semana != "Todas": 
            df_mostrar = df_mostrar[df_mostrar['SEMANA'] == f_semana]
            filtros_usados_c.append(f"Semana: {f_semana}")
            
        if f_busqueda_compras:
            mask = df_mostrar.astype(str).apply(lambda x: x.str.contains(f_busqueda_compras, case=False)).any(axis=1)
            df_mostrar = df_mostrar[mask]
            filtros_usados_c.append(f"Búsqueda: '{f_busqueda_compras}'")
            
        texto_f = " | ".join(filtros_usados_c) if filtros_usados_c else "Ninguno (Mostrando todo)"
        
        if not df_mostrar.empty:
            cols_ordenadas = ['ID_REQ', 'TIPO_REQ', 'FECHA_SOLICITUD', 'FECHA_COMPRA', 'ZONA', 'UNIDAD', 'SOLICITANTE', 'CANTIDAD', 'DESCRIPCION_ITEM', 'TIEMPO_SLA', 'ESTADO', 'NOTA_SOLICITUD', 'NOTA_COMPRA', 'VALOR', 'FACTURA']
            cols_existentes = [c for c in cols_ordenadas if c in df_mostrar.columns]
            df_mostrar_ordenado = df_mostrar[cols_existentes]

            def estilo_tabla_html(row):
                styles = [''] * len(row)
                for i, col in enumerate(row.index):
                    base_style = 'border: 1px solid black; text-align: center; padding: 10px; font-size: 13px; color: black; '
                    if col == 'ESTADO':
                        val = str(row[col]).upper()
                        if 'PENDIENTE' in val: styles[i] = base_style + 'background-color: #FFEFEF; color: #DC3545; font-weight: bold;'
                        elif 'PARCIAL' in val: styles[i] = base_style + 'background-color: #FFF3CD; color: #856404; font-weight: bold;'
                        elif 'COMPRADO' in val: styles[i] = base_style + 'background-color: #E8F5E9; color: #198754; font-weight: bold;'
                        else: styles[i] = base_style
                    elif col == 'TIEMPO_SLA':
                        val = str(row[col])
                        if '⚠️' in val: styles[i] = base_style + 'background-color: #FFEFEF; color: #DC3545; font-weight: bold;'
                        elif '⏳' in val: styles[i] = base_style + 'color: #856404; font-weight: bold;'
                        elif '✅' in val: styles[i] = base_style + 'color: #198754; font-weight: bold;'
                        else: styles[i] = base_style
                    elif col in ['DESCRIPCION_ITEM', 'NOTA_SOLICITUD', 'NOTA_COMPRA', 'TIPO_REQ']:
                        styles[i] = base_style + 'text-align: left;'
                    else:
                        styles[i] = base_style
                return styles

            estilos_html_css = [
                dict(selector="table", props=[("width", "100%"), ("border-collapse", "collapse"), ("font-family", "sans-serif"), ("border", "1px solid black"), ("background-color", "white")]),
                dict(selector="thead th", props=[("background-color", "#1A3B5C"), ("color", "white"), ("font-weight", "bold"), ("text-align", "center"), ("padding", "12px"), ("border", "1px solid black"), ("font-size", "14px")]),
                dict(selector="tbody td", props=[("border", "1px solid black")])
            ]
            
            tabla_html = df_mostrar_ordenado.style.apply(estilo_tabla_html, axis=1).set_table_styles(estilos_html_css).hide(axis="index").to_html()
            st.markdown(tabla_html, unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            pdf_historial = crear_pdf_historial(df_mostrar_ordenado, texto_f)
            st.download_button(
                label="📄 Descargar Tabla Actual en PDF",
                data=pdf_historial,
                file_name=f"Historial_Compras_{obtener_hora_venezuela().strftime('%d%m%y')}.pdf",
                mime="application/pdf",
                use_container_width=True
            )
        else:
            st.warning("No hay registros que coincidan con los filtros seleccionados.")

    # --- PESTAÑA 4: GENERADOR DE PLANILLAS MÚLTIPLES ---
    with tab4:
        st.markdown("### 🖨️ Generador de Planilla Oficial (Múltiples Ítems)")
        st.info("Paso 1: Utiliza los filtros para encontrar las solicitudes. Paso 2: Selecciónalas para armar la planilla (Máximo 13).")
        
        if df_req.empty:
            st.warning("No hay solicitudes en la base de datos.")
        else:
            st.markdown("#### 🔍 1. Filtrar Búsqueda")
            col_f1, col_f2, col_f3 = st.columns(3)
            
            fechas_disp = ["Todas"] + sorted(df_req['FECHA_SOLICITUD'].astype(str).unique().tolist(), reverse=True) 
            zonas_disp = ["Todas"] + sorted(df_req['ZONA'].astype(str).unique().tolist())
            unidades_disp = ["Todas"] + sorted(df_req['UNIDAD'].astype(str).unique().tolist())
            
            with col_f1: f_fecha_plan = st.selectbox("📅 Fecha de Solicitud:", fechas_disp, key="fp_fecha")
            with col_f2: f_zona_plan = st.selectbox("📍 Zona Operativa:", zonas_disp, key="fp_zona")
            with col_f3: f_unidad_plan = st.selectbox("🚛 Unidad / Placa:", unidades_disp, key="fp_unidad")
            
            df_filtrado_plan = df_req.copy()
            if f_fecha_plan != "Todas": df_filtrado_plan = df_filtrado_plan[df_filtrado_plan['FECHA_SOLICITUD'] == f_fecha_plan]
            if f_zona_plan != "Todas": df_filtrado_plan = df_filtrado_plan[df_filtrado_plan['ZONA'] == f_zona_plan]
            if f_unidad_plan != "Todas": df_filtrado_plan = df_filtrado_plan[df_filtrado_plan['UNIDAD'] == f_unidad_plan]
            
            st.markdown("---")
            st.markdown("#### ✅ 2. Seleccionar Ítems")
            
            if df_filtrado_plan.empty:
                st.warning("No hay solicitudes que coincidan con estos filtros específicos.")
            else:
                opciones_multiples = df_filtrado_plan['ID_REQ'] + " | " + df_filtrado_plan['FECHA_SOLICITUD'] + " | " + df_filtrado_plan['UNIDAD'] + " - " + df_filtrado_plan['DESCRIPCION_ITEM']
                diccionario_tickets = dict(zip(opciones_multiples, df_filtrado_plan['ID_REQ']))
                
                selecciones = st.multiselect(
                    "Selecciona los ítems a imprimir:", 
                    opciones_multiples.tolist(),
                    max_selections=13
                )
                
                if selecciones:
                    st.success(f"Has seleccionado {len(selecciones)} ítems para la planilla.")
                    col_b1, col_b2 = st.columns(2)
                    
                    ids_a_buscar = [diccionario_tickets[s] for s in selecciones]
                    df_final_imprimir = df_req[df_req['ID_REQ'].isin(ids_a_buscar)] 
                    
                    notas_completas = []
                    for _, fila in df_final_imprimir.iterrows():
                        nota_sol = str(fila.get('NOTA_SOLICITUD', '')).strip()
                        nota_com = str(fila.get('NOTA_COMPRA', '')).strip()
                        if nota_sol: notas_completas.append(f"[{fila['ID_REQ']}] Solicitud: {nota_sol}")
                        if nota_com: notas_completas.append(f"[{fila['ID_REQ']}] Compra: {nota_com}")
                    texto_notas_final = " | ".join(notas_completas) if notas_completas else "Sin observaciones adicionales."
                    
                    with col_b1:
                        if st.button("📥 Generar Planilla en EXCEL", use_container_width=True):
                            try:
                                wb = openpyxl.load_workbook("Planilla de Solicitud de compra.xlsx")
                                ws = wb.active
                                
                                for i, (_, row_data) in enumerate(df_final_imprimir.iterrows()):
                                    fila_excel = 8 + i
                                    ws[f'B{fila_excel}'] = i + 1
                                    ws[f'C{fila_excel}'] = str(row_data.get('CANTIDAD', ''))
                                    ws[f'D{fila_excel}'] = str(row_data.get('DESCRIPCION_ITEM', ''))
                                
                                ws['B22'] = texto_notas_final
                                output = io.BytesIO()
                                wb.save(output)
                                output.seek(0)
                                
                                st.download_button(
                                    label="⬇️ Descargar Archivo Excel",
                                    data=output,
                                    file_name=f"Planilla_Compras_{obtener_hora_venezuela().strftime('%d%m%y')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    type="primary"
                                )
                            except FileNotFoundError:
                                st.error("⚠️ Falta el archivo 'Planilla de Solicitud de compra.xlsx' en tu carpeta.")
                                
                    with col_b2:
                        pdf_oficial = crear_pdf_planilla_oficial(df_final_imprimir, texto_notas_final)
                        st.download_button(
                            label="📄 Descargar Planilla en PDF",
                            data=pdf_oficial,
                            file_name=f"Planilla_Oficial_{obtener_hora_venezuela().strftime('%d%m%y')}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            type="primary"
                        )